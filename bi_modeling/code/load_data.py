import json
import os
from pathlib import Path
import pandas as pd
import numpy as np

DATA_DIR_PATH = Path(__file__).resolve().parents[1] / "data"
DATA_DIR = os.fspath(DATA_DIR_PATH)
DISTRIBUTION_DATA_DIR_PATH = DATA_DIR_PATH / "distribution"
RECEIVING_AREA_SHARE_DATA_PATH = DISTRIBUTION_DATA_DIR_PATH / "receiving_area_shares.json"
BII_DATA_DIR_PATH = DATA_DIR_PATH / "BII"
BII_DISTRIBUTION_LOOKUP_PATH = BII_DATA_DIR_PATH / "bii_2023_distribution_lookup.json"
MIDPOINT_TO_ENDPOINT_WORKBOOK_PATH = DATA_DIR_PATH / "ReCiPe2016_CFs_v1.1_20180117.xlsx"
RECIPE_COUNTRY_CF_PATH = DATA_DIR_PATH / "ReCiPe2016_country factors_v1.1_20171221.xlsx"
MIDPOINT_TO_ENDPOINT_SHEET_NAME = "Midpoint to endpoint factors"
WATER_CONSUMPTION_COUNTRY_SHEET_NAME = "Water consumption"
DEFAULT_RECIPE_PERSPECTIVE = "hierarchist"
RECIPE_PERSPECTIVE_ALIASES = {
    "individualistic": "individualistic",
    "individualist": "individualistic",
    "hierarchic": "hierarchist",
    "hierarchist": "hierarchist",
    "egalitarian": "egalitarian",
}
ENDPOINT_SCOPE_ORDER = ("human_health", "terrestrial", "freshwater", "marine", "resources")
ECOSYSTEM_ENDPOINT_SCOPES = ("terrestrial", "freshwater", "marine")
MIDPOINT_TO_ENDPOINT_ROW_MAP = {
    "Global Warming - Human health": ("human_health", "GWP"),
    "Stratospheric ozone depletion - Human health": ("human_health", "ODP"),
    "Ionzing Radiation - Human health": ("human_health", "IR"),
    "Fine particulate matter formation - Human health": ("human_health", "PMFP"),
    "Photochemical ozone formation - Human health": ("human_health", "POFP"),
    "Toxicity - Human health (cancer)": ("human_health", "HTPc"),
    "Toxicity - Human health (non-cancer)": ("human_health", "HTPnc"),
    "Water consumption - human health": ("human_health", "WC"),
    "Global Warming - Terrestrial ecosystems": ("terrestrial", "GWP"),
    "Photochemical ozone formation - Terrestrial ecosystems": ("terrestrial", "POFP"),
    "Acidification - Terrestrial ecosystems": ("terrestrial", "AP"),
    "Toxicity - Terrestrial ecosystems": ("terrestrial", "TETP"),
    "Water consumption - terrestrial ecosystems": ("terrestrial", "WC"),
    "Land use - occupation and transformation": ("terrestrial", "LU"),
    "Global Warming - Freshwater ecosystems": ("freshwater", "GWP"),
    "Eutrophication - Freshwater ecosystems": ("freshwater", "FEP"),
    "Toxicity - Freshwater ecosystems": ("freshwater", "FETP"),
    "Water consumption -aquatic ecosystems": ("freshwater", "WC"),
    "Toxicity - Marine ecosystems": ("marine", "METP"),
    "Eutrophication - Marine ecosystems": ("marine", "MEP"),
    "Mineral resource scarcity": ("resources", "MRS"),
}


def _empty_midpoint_to_endpoint_scope_map():
    return {scope: {} for scope in ENDPOINT_SCOPE_ORDER}


def _build_hierarchist_midpoint_to_endpoint_fallback():
    fallback = _empty_midpoint_to_endpoint_scope_map()
    ecosystem_factors = cf_data["impact_factors"]["midpoint_to_endpoint"]
    for scope in ECOSYSTEM_ENDPOINT_SCOPES:
        if scope in ecosystem_factors:
            fallback[scope] = {
                midpoint: float(value)
                for midpoint, value in ecosystem_factors[scope].items()
                if isinstance(value, (int, float))
            }
    return fallback


def load_midpoint_to_endpoint_factors(file_path=MIDPOINT_TO_ENDPOINT_WORKBOOK_PATH):
    """
    Load perspective-aware midpoint-to-endpoint factors from the ReCiPe 2016
    workbook. Keys are normalized to ``individualistic``, ``hierarchist``,
    and ``egalitarian``.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[MIDPOINT_TO_ENDPOINT_SHEET_NAME]
        header = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
        perspective_columns = {}
        for column_idx, header_value in enumerate(header, start=1):
            if header_value is None:
                continue
            normalized = RECIPE_PERSPECTIVE_ALIASES.get(str(header_value).strip().lower())
            if normalized is not None:
                perspective_columns[normalized] = column_idx

        if not perspective_columns:
            raise ValueError(
                f"No recognized perspective columns were found in {file_path} "
                f"sheet {MIDPOINT_TO_ENDPOINT_SHEET_NAME!r}."
            )

        midpoint_to_endpoint_by_perspective = {
            perspective: _empty_midpoint_to_endpoint_scope_map()
            for perspective in perspective_columns
        }

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            row_label = row[0]
            if row_label not in MIDPOINT_TO_ENDPOINT_ROW_MAP:
                continue
            scope, midpoint = MIDPOINT_TO_ENDPOINT_ROW_MAP[row_label]
            for perspective, column_idx in perspective_columns.items():
                value = row[column_idx - 1]
                if isinstance(value, (int, float)):
                    midpoint_to_endpoint_by_perspective[perspective][scope][midpoint] = float(value)
    finally:
        workbook.close()

    return midpoint_to_endpoint_by_perspective


def load_location_specific_water_endpoint_factors(file_path=RECIPE_COUNTRY_CF_PATH):
    """
    Load country-specific ReCiPe 2016 endpoint factors for water consumption.

    The workbook provides terrestrial ecosystem factors by perspective and a
    freshwater/aquatic ecosystem factor shared across perspectives.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[WATER_CONSUMPTION_COUNTRY_SHEET_NAME]
        factors = {
            "terrestrial": {},
            "freshwater": {},
        }

        for row in worksheet.iter_rows(min_row=7, values_only=True):
            if not row or len(row) < 10:
                continue
            country = row[5]
            if country is None:
                continue
            country_name = str(country).strip()
            if not country_name:
                continue

            terrestrial_values = {
                "individualistic": row[6] if isinstance(row[6], (int, float)) else None,
                "hierarchist": row[7] if isinstance(row[7], (int, float)) else None,
                "egalitarian": row[8] if isinstance(row[8], (int, float)) else None,
            }
            if any(value is not None for value in terrestrial_values.values()):
                factors["terrestrial"][country_name] = terrestrial_values

            freshwater_value = row[9] if isinstance(row[9], (int, float)) else None
            if freshwater_value is not None:
                factors["freshwater"][country_name] = {
                    perspective: float(freshwater_value)
                    for perspective in RECIPE_PERSPECTIVE_ALIASES.values()
                }
    finally:
        workbook.close()

    return factors

def load_json_file(file_path):
    """
    Load a JSON file and return its content.
    
    :param file_path: Path to the JSON file.
    :return: Parsed JSON content as a dictionary.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"The file {file_path} does not exist.")
    
    with open(file_path, 'r', encoding='utf-8') as file:
        data = json.load(file)
    
    return data


def load_receiving_area_share_data(file_path=RECEIVING_AREA_SHARE_DATA_PATH):
    """
    Load the processed receiving-area dataset used by the distribution layer.
    """
    return load_json_file(os.fspath(file_path))


def load_bii_distribution_data(file_path=BII_DISTRIBUTION_LOOKUP_PATH):
    """
    Load the processed BII lookup used by the distribution layer.
    """
    return load_json_file(os.fspath(file_path))

cf_data = load_json_file(os.path.join(DATA_DIR, "characterization_factors.json"))
ef_data = load_json_file(os.path.join(DATA_DIR, "energy_and_emission_factors/grid_emission_factors_2016_2024.json"))

impact_factors = cf_data['impact_factors']['manufacturing']
TO_CTUe_CONVERSION_FACTOR_FRESHWATER = 983.513372389555 # 1,4‑DCB emitted to continental freshwater, TRACI 2.2 
TO_CTUe_CONVERSION_FACTOR_MARINE = 0.802271239202204 # 1,4‑DCB emitted to marine water, TRACI 2.2
# unified_emission_factors = cf_data['impact_factors']['emission_factors']
unified_emission_factors = ef_data['emission_factors']
cf_transportation = cf_data['impact_factors']['transportation']
cf_eol = cf_data['impact_factors']['eol_factors']
cf_bom = cf_data['impact_factors']['raw_material_factors']
try:
    midpoint_to_endpoint_factors_by_perspective = load_midpoint_to_endpoint_factors()
except Exception:
    midpoint_to_endpoint_factors_by_perspective = {
        DEFAULT_RECIPE_PERSPECTIVE: _build_hierarchist_midpoint_to_endpoint_fallback()
    }
midpoint_to_endpoint_factors = midpoint_to_endpoint_factors_by_perspective[DEFAULT_RECIPE_PERSPECTIVE]
try:
    location_specific_water_endpoint_factors = load_location_specific_water_endpoint_factors()
except Exception:
    location_specific_water_endpoint_factors = {
        "terrestrial": {},
        "freshwater": {},
    }
try:
    receiving_area_distribution_data = load_receiving_area_share_data()
except Exception:
    receiving_area_distribution_data = {
        "metadata": {},
        "areas": {},
        "region_area_lookup": {},
    }
try:
    bii_distribution_data = load_bii_distribution_data()
except Exception:
    bii_distribution_data = {
        "metadata": {},
        "area_bii_2023": {},
        "bucket_default_bii_2023": {},
    }

me_data = load_json_file(os.path.join(DATA_DIR, "manufacturing_emissions.json"))

tsmc_acid_emissions_raw_pixels = me_data['tsmc_acid_emissions_raw_pixels']
tsmc_acid_emission_mix_ratio = {year:{} for year in tsmc_acid_emissions_raw_pixels.keys()}
for year in tsmc_acid_emissions_raw_pixels.keys():
    total_acid = tsmc_acid_emissions_raw_pixels[year]['total']
    for acid in tsmc_acid_emissions_raw_pixels[year].keys():
        if acid != 'total':
            tsmc_acid_emission_mix_ratio[year][acid] = tsmc_acid_emissions_raw_pixels[year][acid] / total_acid

tsmc_emissions_macro = me_data['tsmc_emissions_macro']
tsmc_electricity_consumption = me_data['tsmc_electricity_consumption']
node_to_layer_masks_map = me_data['node_to_layer_masks']
embodied_carbon_and_water = me_data['carbon_and_water_intensity_by_node']
dram_year_to_node_map = me_data['dram_year_to_node_map']
ssd_year_to_node_map = me_data['ssd_year_to_node_map']

hynix_emissions_macro = me_data['hynix_emissions_macro']
hynix_production_data = me_data['hynix_production_data']
spil_packaging_emissions = me_data['spil_packaging_emissions']
spil_testing_emissions = me_data['spil_testing_emissions']

# Create dictionaries to store the allocated emissions
dram_emissions = {}
nand_emissions = {}

# List of emission types to process
emission_types = ['SOx', 'NOx', 'VOC', 'HF', 'HCl', 'NH3', 'total_wastewater_nitrogen', 
                 'total_wastewater_phosphorus', 'wastewater_fluoride', 'ammonia_nitrogen']

# Calculate allocated emissions for each year and type
for year_idx, year in enumerate(range(2016, 2025)):
    dram_emissions[year] = {}
    nand_emissions[year] = {}
    
    dram_ratio = hynix_production_data['dram_revenue_ratio'][year_idx]
    nand_ratio = hynix_production_data['nand_revenue_ratio'][year_idx]
    
    dram_wafers = hynix_production_data['estimated_dram_k_wafers'][year_idx] * 1000  # Convert to wafers
    nand_wafers = hynix_production_data['nand_k_wafers'][year_idx] * 1000  # Convert to wafers

    wastewater_per_wafer_dram = hynix_emissions_macro['wastewater_discharge_1000_m3'][year_idx] * dram_ratio * 1e6 / dram_wafers  # L/wafer
    wastewater_per_wafer_nand = hynix_emissions_macro['wastewater_discharge_1000_m3'][year_idx] * nand_ratio * 1e6 / nand_wafers  # L/wafer
    dram_emissions[year]['Cu2+'] = wastewater_per_wafer_dram * tsmc_emissions_macro['Cu2+_ppm'][year_idx] * 1e-6  # kg/wafer
    nand_emissions[year]['Cu2+'] = wastewater_per_wafer_nand * tsmc_emissions_macro['Cu2+_ppm'][year_idx] * 1e-6  # kg/wafer
    
    for emission_type in emission_types:
        total_emission = hynix_emissions_macro[f'{emission_type}_mt'][year_idx]* 1000 # Convert to kg
        
        # Calculate allocated emissions   
        dram_emissions[year][emission_type] = (total_emission * dram_ratio ) / dram_wafers
        nand_emissions[year][emission_type] = (total_emission * nand_ratio ) / nand_wafers

dram_bit_density_by_year = me_data['dram_bit_density_by_year']
vram_bit_density = me_data['vram_bit_density']
ssd_bit_density_by_year = me_data['ssd_bit_density_by_year']

ds_data = load_json_file(os.path.join(DATA_DIR, "device_specs.json"))
regional_direct_wue_data = load_json_file(os.path.join(DATA_DIR, "wet_bulb_temperature/regional_direct_wue_compact.json"))
regional_device_transport_routes = load_json_file(os.path.join(DATA_DIR, "transportation/regional_device_transport_routes.json"))
missing_transport_regions = sorted(
    set(regional_direct_wue_data["regions"].keys()) - set(regional_device_transport_routes["regions"].keys())
)
extra_transport_regions = sorted(
    set(regional_device_transport_routes["regions"].keys()) - set(regional_direct_wue_data["regions"].keys())
)
if missing_transport_regions or extra_transport_regions:
    raise ValueError(
        "Transport route regions must stay aligned with the regional direct-WUE dataset. "
        f"Missing transport regions: {missing_transport_regions}; extra transport regions: {extra_transport_regions}"
    )

epyc_7b12_specs = ds_data['CPUs']['EPYC 7B12']
epyc_7443_specs = ds_data['CPUs']['EPYC 7443']
epyc_7b13_specs = ds_data['CPUs']['EPYC 7B13']
epyc_9b14_specs = ds_data['CPUs']['EPYC 9B14']
epyc_9b45_specs = ds_data['CPUs']['EPYC 9B45']

t4_specs = ds_data['GPUs']['T4']
v100_specs = ds_data['GPUs']['V100']
l40_specs = ds_data['GPUs']['L40']
a100_40g_specs = ds_data['GPUs']['A100 40GB']
h100_specs = ds_data['GPUs']['H100']

bom_templates = ds_data['BoM_template']

df_storage_weight_density = pd.DataFrame.from_dict(ds_data['storage_weight_density'])

if os.path.exists(os.path.join(DATA_DIR, "hdd_impact_factors_by_stage_year.json")):
    with open(os.path.join(DATA_DIR, "hdd_impact_factors_by_stage_year.json"), 'r', encoding='utf-8') as f:
        hdd_impact_factors_by_stage_year = json.load(f)
