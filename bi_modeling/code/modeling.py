import numpy as np
import pandas as pd
from functools import lru_cache
from pathlib import Path
from sklearn.linear_model import LinearRegression

from load_data import *

DRAM_SSD_PACKAGING_WEIGHT_FACTOR = 2.0
CPU_GPU_PACKAGING_WEIGHT_FACTOR = 1.2
PO4_TO_P_CONVERSION_FACTOR = 0.33
# EFFECTIVE_WAFER_AREA_MM2 = np.pi * (300 / 2 - 2)**2  # 300mm wafer with 2mm edge exclusion
EFFECTIVE_WAFER_AREA_MM2 = 440 * 100 # Bsed on the CO2 per wafer and per cm2 values reported by IEDM 23 paper Cradle-to-gate Life Cycle Assessment of CMOS Logic Technologies, which implies an effective area of around 440 cm2 for a 300mm wafer after accounting for edge exclusion and other non-productive areas.
# The most likely hypothesis is: the square inscribed in a circle with diameter 300mm has a side length of 300/sqrt(2) ~ 212mm, giving an area of around 44900 mm2 or 449 cm2. The additional reduction to 440 cm2 could be due to further exclusions for non-productive areas, test structures, scribe lines, or other factors that reduce the effective area available for die production.
MIDPOINT_SCOPES_DEFAULT = ['AP', 'FEP', 'MEP', 'POFP', 'TETP', 'FETP', 'METP', 'GWP', 'WC']
MIDPOINT_SCOPES = list(MIDPOINT_SCOPES_DEFAULT)


def get_midpoint_scopes():
    """Return the active midpoint scopes tuple used by modeling helpers."""
    return tuple(MIDPOINT_SCOPES)


def set_midpoint_scopes(midpoint_scopes):
    """Override active midpoint scopes for runtime experiments.

    The override is process-local and affects helpers that iterate the
    module-level ``MIDPOINT_SCOPES`` list.
    """
    global MIDPOINT_SCOPES
    if midpoint_scopes is None:
        MIDPOINT_SCOPES = list(MIDPOINT_SCOPES_DEFAULT)
        return tuple(MIDPOINT_SCOPES)

    normalized = []
    for scope in midpoint_scopes:
        scope_key = str(scope).strip()
        if not scope_key:
            continue
        if scope_key in normalized:
            continue
        normalized.append(scope_key)

    if not normalized:
        raise ValueError("midpoint_scopes override must contain at least one scope.")

    MIDPOINT_SCOPES = normalized
    return tuple(MIDPOINT_SCOPES)


def reset_midpoint_scopes():
    """Reset active midpoint scopes to repository defaults."""
    return set_midpoint_scopes(MIDPOINT_SCOPES_DEFAULT)

TRANSPORT_DISTANCE_KEYS = ("Truck", "Ship", "Air")
LOGIC_COMPONENT_TYPES = {"CPU", "GPU"}
MEMORY_STORAGE_COMPONENT_TYPES = {"DRAM", "SSD", "HDD"}
MANUFACTURING_SCOPE_KEYS = ("scope1", "scope2", "scope3")
IMPACT_FACTOR_POLLUTANT_ALIASES = {
    'CO2e': ('CO2',),
    'P-total_wastewater_phosphorus': ('total_wastewater_phosphorus',),
    'PO43-': ('phosphate', 'PO_4^{3-}'),
    'PO43- - phosphate': ('phosphate', 'PO_4^{3-}'),
}
EMISSION_FACTOR_LOCATION_ALIASES = {
    'US-CA': 'CA',
    'US-CAL-CISO': 'CA',
    'US-TX': 'TX',
    'US-TEX-ERCO': 'TX',
    'US-VA': 'VA',
    'US-MIDA-PJM': 'VA',
    'US-MIDW-MISO': 'MidWest',
    'Taiwan_EDGAR': 'Taiwan_g/kWh_EDGAR',
    'Korea_EDGAR': 'Korea_g/kWh_EDGAR',
}
SPIL_BACKEND_DIRECT_POLLUTANT_MAP = {
    'voc_emissions_kg_per_packaged_ic': 'VOC',
    'effluent_ammonia_nitrogen_kg_per_packaged_ic': 'ammonia_nitrogen',
    'effluent_cod_kg_per_packaged_ic': 'COD',
    'effluent_copper_kg_per_packaged_ic': 'Cu2+',
    'effluent_fluoride_salt_kg_per_packaged_ic': 'wastewater_fluoride',
    'effluent_iron_kg_per_packaged_ic': 'Fe2+',
    'effluent_lead_kg_per_packaged_ic': 'Pb2+',
    'effluent_nickel_kg_per_packaged_ic': 'Ni2+',
}
RECIPE_COUNTRY_CF_PATH = Path(__file__).resolve().parents[1] / "data" / "ReCiPe2016_country factors_v1.1_20171221.xlsx"
LOCATION_SPECIFIC_WORKBOOK_CONFIG = {
    'AP': {
        'sheet': 'Terrestrial acidification',
        'columns': {'NOx': 2, 'NH3': 3, 'SOx': 4},  # Workbook column is SO2; use SOx key for model compatibility.
        'aliases': {
            'Spain': ('Spain',),
            'Germany': ('Germany',),
            'France': ('France',),
            'US': ('USA', 'United States'),
            'China': ('China',),
            'Taiwan': ('Taiwan', 'China'),
            'Korea': ('South Korea',),
        },
    },
    'FEP': {
        'sheet': 'Freshwater eutrophication',
        'columns': {
            'total_wastewater_phosphorus': 2,
            'P-total_wastewater_phosphorus': 2,
            'phosphate': 3,
            'PO_4^{3-}': 3,
            'PO43-': 3,
            'PO43- - phosphate': 3,
        },
        'aliases': {
            'Spain': ('Spain',),
            'Germany': ('Germany',),
            'France': ('France',),
            'US': ('United States', 'USA'),
            'China': ('China',),
            'Taiwan': ('Taiwan', 'China'),
            'Korea': ('South Korea',),
        },
    },
    'POFP': {
        'sheet': 'Photochemical ozone formation',
        'columns': {'NOx': 5, 'NMVOC': 6, 'VOC': 6},  # Ecosystem Ozone Formation Potential (EOFP), matching the legacy overrides.
        'aliases': {
            'Spain': ('Spain, Portugal', 'Spain'),
            'Germany': ('Germany',),
            'France': ('France, Andorra', 'France'),
            'US': ('United States', 'USA'),
            'China': ('China, Hong Kong, Macao', 'China'),
            'Taiwan': ('Taiwan',),
            'Korea': ('South Korea',),
        },
    },
}
LOGIC_SCOPE_DENSITY_KEYS = {
    'scope1': 'carbon_density_scope_1_cmos_kgCO2e_per_cm2_by_node',
    'scope2': 'carbon_density_scope_2_cmos_kgCO2e_per_cm2_by_node',
    'scope3': 'carbon_density_scope_3_cmos_kgCO2e_per_cm2_by_node',
}
LOGIC_SCOPE_NODE_ALIASES = {
    '12': '14',
}
STORAGE_SCOPE_PER_WAFER_KEYS = {
    'DRAM': {
        'scope1': 'dram_carbon_emissions_scope1_kgCO2e_per_wafer_by_node',
        'scope2': 'dram_carbon_emissions_scope2_kgCO2e_per_wafer_by_node',
    },
    'HBM': {
        'scope1': 'dram_carbon_emissions_scope1_kgCO2e_per_wafer_by_node',
        'scope2': 'dram_carbon_emissions_scope2_kgCO2e_per_wafer_by_node',
    },
    'SSD': {
        'scope1': 'ssd_carbon_emission_scope1_kgCO2e_per_wafer_by_node',
        'scope2': 'ssd_carbon_emission_scope2_kgCO2e_per_wafer_by_node',
    },
}


@lru_cache(maxsize=1)
def _load_location_specific_midpoint_factors():
    """
    Load the supported location-specific midpoint CF overrides directly from the
    ReCiPe 2016 workbook and cache them for repeated operational-impact calls.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required to load location-specific midpoint factors from "
            f"{RECIPE_COUNTRY_CF_PATH}. Run with ~\\miniconda3\\python.exe or install openpyxl."
        ) from exc

    workbook = load_workbook(RECIPE_COUNTRY_CF_PATH, read_only=True, data_only=True)
    midpoint_factors = {midpoint: {} for midpoint in LOCATION_SPECIFIC_WORKBOOK_CONFIG}

    try:
        for midpoint, config in LOCATION_SPECIFIC_WORKBOOK_CONFIG.items():
            sheet = workbook[config['sheet']]
            row_lookup = {}
            for row in sheet.iter_rows(values_only=True):
                if not row or row[0] is None:
                    continue
                region_name = str(row[0]).strip()
                if region_name:
                    row_lookup[region_name] = row

            for location, aliases in config['aliases'].items():
                matched_row = None
                for alias in aliases:
                    if alias in row_lookup:
                        matched_row = row_lookup[alias]
                        break
                if matched_row is None:
                    continue

                midpoint_factors[midpoint][location] = {
                    pollutant: float(matched_row[column_idx - 1])
                    for pollutant, column_idx in config['columns'].items()
                }
    finally:
        workbook.close()

    return midpoint_factors


def _iter_impact_factor_keys(pollutant):
    yield pollutant
    aliases = IMPACT_FACTOR_POLLUTANT_ALIASES.get(pollutant, ())
    if isinstance(aliases, str):
        aliases = (aliases,)
    for alias in aliases:
        if alias != pollutant:
            yield alias


def _normalize_emission_factor_location(location):
    location_str = str(location).strip()
    if not location_str:
        return location_str

    if location_str in EMISSION_FACTOR_LOCATION_ALIASES:
        return EMISSION_FACTOR_LOCATION_ALIASES[location_str]

    uppercase_location = location_str.upper()
    if uppercase_location in EMISSION_FACTOR_LOCATION_ALIASES:
        return EMISSION_FACTOR_LOCATION_ALIASES[uppercase_location]

    if uppercase_location.startswith('US-') and len(uppercase_location) == 5 and uppercase_location[3:].isalpha():
        return uppercase_location[3:]

    return location_str


def _get_edgar_parent_location(location):
    location_str = str(location).strip()
    if not location_str:
        return None
    if location_str.endswith('_g/kWh_EDGAR'):
        return location_str[:-len('_g/kWh_EDGAR')]
    if location_str.endswith('_EDGAR'):
        return location_str[:-len('_EDGAR')]
    return None


def _get_location_specific_midpoint_lookup_candidates(location):
    candidates = []
    normalized_location = _normalize_emission_factor_location(location)

    for candidate in (str(location).strip(), normalized_location):
        if candidate and candidate not in candidates:
            candidates.append(candidate)

    if str(location).strip().upper().startswith('US-') and 'US' not in candidates:
        candidates.append('US')

    return tuple(candidates)


def _get_location_specific_midpoint_cfs(location, spatial_awareness=False):
    if not spatial_awareness:
        return {}

    location_candidates = _get_location_specific_midpoint_lookup_candidates(location)
    midpoint_factors = {}

    for midpoint, factors_by_location in _load_location_specific_midpoint_factors().items():
        for candidate in location_candidates:
            if candidate in factors_by_location:
                midpoint_factors[midpoint] = factors_by_location[candidate]
                break

    return midpoint_factors


def _get_location_specific_water_endpoint_lookup_candidates(location):
    candidates = []
    normalized_location = _normalize_emission_factor_location(location)

    for candidate in (str(location).strip(), normalized_location):
        if candidate and candidate not in candidates:
            candidates.append(candidate)

    alias_key = str(normalized_location).strip().upper()
    for alias in LOCATION_SPECIFIC_WATER_ENDPOINT_ALIASES.get(alias_key, ()):
        if alias not in candidates:
            candidates.append(alias)

    if str(location).strip().upper().startswith('US-'):
        for alias in LOCATION_SPECIFIC_WATER_ENDPOINT_ALIASES['US']:
            if alias not in candidates:
                candidates.append(alias)

    return tuple(candidates)


def _get_location_specific_water_endpoint_factors(location, perspective_key, spatial_awareness=False):
    if not spatial_awareness or not location:
        return {}

    selected_factors = {}
    lookup_candidates = _get_location_specific_water_endpoint_lookup_candidates(location)

    for scope in ('terrestrial', 'freshwater'):
        factors_by_country = location_specific_water_endpoint_factors.get(scope, {})
        for candidate in lookup_candidates:
            if candidate not in factors_by_country:
                continue
            factor = factors_by_country[candidate].get(perspective_key)
            if factor is not None:
                selected_factors[scope] = float(factor)
                break

    return selected_factors


def _get_midpoint_impact_factor(impact_type, pollutant, location_specific_cfs=None):
    candidate_keys = tuple(_iter_impact_factor_keys(pollutant))
    location_specific_lookup = (location_specific_cfs or {}).get(impact_type, {})
    midpoint_factor_lookup = impact_factors.get(impact_type, {})

    for candidate_key in candidate_keys:
        if candidate_key in location_specific_lookup:
            return location_specific_lookup[candidate_key]

    if not isinstance(midpoint_factor_lookup, dict):
        return None

    for candidate_key in candidate_keys:
        if candidate_key in midpoint_factor_lookup:
            return midpoint_factor_lookup[candidate_key]

    return None


def _aggregate_midpoint_group_impacts(grouped_impacts):
    totals = {midpoint: 0.0 for midpoint in MIDPOINT_SCOPES}
    if not isinstance(grouped_impacts, dict):
        return totals

    if any(midpoint in grouped_impacts for midpoint in MIDPOINT_SCOPES):
        for midpoint in MIDPOINT_SCOPES:
            totals[midpoint] += grouped_impacts.get(midpoint, 0.0)
        return totals

    for impact_group in grouped_impacts.values():
        if isinstance(impact_group, dict):
            for midpoint in MIDPOINT_SCOPES:
                totals[midpoint] += impact_group.get(midpoint, 0.0)

    return totals


def _empty_midpoint_impacts(midpoint_scopes=None):
    return {midpoint: 0.0 for midpoint in (midpoint_scopes or MIDPOINT_SCOPES)}


def _empty_manufacturing_scope_breakdown(midpoint_scopes=None):
    selected_midpoints = tuple(midpoint_scopes or MIDPOINT_SCOPES)
    return {
        scope: _empty_midpoint_impacts(selected_midpoints)
        for scope in MANUFACTURING_SCOPE_KEYS
    }


def _add_midpoint_impacts(target, source, midpoint_scopes=None):
    selected_midpoints = tuple(midpoint_scopes or MIDPOINT_SCOPES)
    if not source:
        return target
    for midpoint in selected_midpoints:
        target[midpoint] += source.get(midpoint, 0.0)
    return target


def _sum_manufacturing_scope_breakdown(scope_breakdown, midpoint_scopes=None):
    totals = _empty_midpoint_impacts(midpoint_scopes)
    for scope_name in MANUFACTURING_SCOPE_KEYS:
        _add_midpoint_impacts(totals, scope_breakdown.get(scope_name, {}), midpoint_scopes)
    return totals


def _build_impact_payload(midpoint_impacts, perspective):
    return {
        'midpoint': dict(midpoint_impacts),
        'endpoint': midpoint_to_endpoint(midpoint_impacts, perspective=perspective),
    }


def _build_scope_breakdown_payload(scope_breakdown, perspective):
    return {
        scope_name: _build_impact_payload(scope_breakdown[scope_name], perspective)
        for scope_name in MANUFACTURING_SCOPE_KEYS
    }


def _material_cf_contribution(material_name, midpoint, mass_kg):
    material_factors = cf_bom.get(material_name)
    if material_factors is None:
        return 0.0
    factor = material_factors.get(midpoint)
    if factor is None:
        return 0.0
    reference_weight_kg = material_factors.get('reference_weight_kg', 1) or 1
    return factor * mass_kg / reference_weight_kg


def _normalize_logic_scope_node(node_nm):
    node_key = str(node_nm)
    return LOGIC_SCOPE_NODE_ALIASES.get(node_key, node_key)


def _get_logic_scope_density(scope_name, node_nm):
    node_key = _normalize_logic_scope_node(node_nm)
    density_key = LOGIC_SCOPE_DENSITY_KEYS[scope_name]
    density = embodied_carbon_and_water[density_key].get(node_key)
    return 0.0 if density is None else float(density)


def _get_logic_total_density(node_nm):
    node_key = _normalize_logic_scope_node(node_nm)
    density = embodied_carbon_and_water['carbon_density_cmos_kgCO2e_per_cm2_by_node'].get(node_key)
    return 0.0 if density is None else float(density)


def _estimate_logic_fab_carbon_scope_split(cpu_specs, production_yield=0.875):
    scope_totals = {scope_name: 0.0 for scope_name in MANUFACTURING_SCOPE_KEYS}
    estimated_total = 0.0
    die_regions = (
        ('die_size_mm2', 'technology_node_nm'),
        ('io_die_size_mm2', 'io_die_technology_node_nm'),
    )

    for area_key, node_key in die_regions:
        area_mm2 = cpu_specs.get(area_key, 0) or 0
        node_nm = cpu_specs.get(node_key)
        if area_mm2 <= 0 or node_nm in (None, '', 0):
            continue

        total_density = _get_logic_total_density(node_nm)
        scope_densities = {
            scope_name: _get_logic_scope_density(scope_name, node_nm)
            for scope_name in MANUFACTURING_SCOPE_KEYS
        }
        scope3_density = scope_densities['scope3']
        if scope3_density == 0.0 and total_density > 0.0:
            scope_densities['scope3'] = max(total_density - scope_densities['scope1'] - scope_densities['scope2'], 0.0)

        for scope_name in MANUFACTURING_SCOPE_KEYS:
            scope_totals[scope_name] += area_mm2 * scope_densities[scope_name] / production_yield / 100

        estimated_total += area_mm2 * total_density / production_yield / 100

    legacy_total = cpu_specs.get('manufacturing_carbon_kgCO2e', 0.0) or 0.0
    if legacy_total == 0.0:
        legacy_total = (
            cpu_specs.get('die_size_mm2', 0)
            * embodied_carbon_and_water['carbon_density_cmos_kgCO2e_per_cm2_by_node'].get(str(cpu_specs['technology_node_nm']), 0)
            / production_yield
            / 100
        )

    if estimated_total > 0.0 and legacy_total > 0.0:
        scale = legacy_total / estimated_total
        for scope_name in MANUFACTURING_SCOPE_KEYS:
            scope_totals[scope_name] *= scale

    return scope_totals


def _get_storage_scope_node(storage_type, production_year, HBM_edition=None):
    if storage_type in ['DRAM', 'HBM']:
        if storage_type == 'HBM':
            return vram_bit_density.get(f'{HBM_edition}_node', '2y')
        return dram_year_to_node_map.get(str(production_year), '')
    return ssd_year_to_node_map.get(str(production_year), '')


def _estimate_storage_fab_carbon_scope_split(storage_type, production_year, equivalent_wafers, HBM_edition=None):
    node = _get_storage_scope_node(storage_type, production_year, HBM_edition=HBM_edition)
    total_key = (
        'dram_carbon_emissions_kgCO2e_per_wafer_by_node'
        if storage_type in ['DRAM', 'HBM']
        else 'ssd_carbon_emissions_kgCO2e_per_wafer_by_node'
    )
    total_per_wafer = embodied_carbon_and_water[total_key].get(node, 0) or 0.0
    scope_totals = {'scope1': 0.0, 'scope2': 0.0, 'scope3': 0.0}

    for scope_name in ('scope1', 'scope2'):
        per_wafer = embodied_carbon_and_water[STORAGE_SCOPE_PER_WAFER_KEYS[storage_type][scope_name]].get(node, 0) or 0.0
        scope_totals[scope_name] = per_wafer * equivalent_wafers

    estimated_total = scope_totals['scope1'] + scope_totals['scope2']
    target_total = total_per_wafer * equivalent_wafers
    if estimated_total > 0.0 and target_total > 0.0:
        scale = target_total / estimated_total
        scope_totals['scope1'] *= scale
        scope_totals['scope2'] *= scale

    return scope_totals

# HDD Modeling -- handled specially in model_hdd_by_stage_impact.py and persisted in hdd_impact_factors_by_stage_year.json

def _get_hdd_stage_impact_factors_for_year(production_year):
    """
    Return HDD impact factors for a production year using the by-stage source loaded in load_data.py.
    Supports both string and integer year keys.
    """
    year_map = globals().get('hdd_impact_factors_by_stage_year')
    if year_map is None:
        from model_hdd_by_stage_impact import hdd_impact_factors_by_stage_year as generated_hdd_stage_year_map
        year_map = generated_hdd_stage_year_map
    if not isinstance(year_map, dict):
        raise ValueError("Invalid HDD by-stage impact factor payload: expected a dictionary.")

    if production_year in year_map:
        return year_map[production_year]
    year_key = str(production_year)
    if year_key in year_map:
        return year_map[year_key]

    available_years = sorted(str(y) for y in year_map.keys())
    raise ValueError(f"HDD production year {production_year} is unavailable in by-stage factors. Available years: {available_years}")

def get_equivalent_wafers(device_specs, bit_density=None, production_yield=0.875):
    device_type = device_specs.get('device_type', device_specs.get('component_type'))
    if device_type is None:
        raise KeyError("device_specs must include either 'device_type' or 'component_type'.")
    if device_type in ['CPU', 'GPU']:
        # Only calculate the compute die for GPU as the HBM is calculated separately with storage manufacturing impact model. This is also to avoid the double counting issue of GPU die and HBM wafer equivalent since they are made by different providers and likely in different regions, which would require different upstream material impact modeling.
        wafer_equivalent = (device_specs['die_size_mm2'] + device_specs.get('io_die_size_mm2', 0)) / EFFECTIVE_WAFER_AREA_MM2 / device_specs.get('production_yield', production_yield)
    elif device_type in ['DRAM', 'SSD', 'HBM']:
        if bit_density is None:
            year_idx = min(max(0, device_specs.get('production_year', 2016) - 2016), 8)  # Ensure year index is within bounds
            if device_type == 'DRAM':
                bit_density = dram_bit_density_by_year[year_idx]
            elif device_type == 'HBM':
                hbm_edition = device_specs.get('hbm_type', 'HBM2')
                bit_density = vram_bit_density[hbm_edition]
            else:  # SSD
                bit_density = ssd_bit_density_by_year[year_idx]
        wafer_equivalent = (device_specs['capacity'] * 8) / (bit_density * EFFECTIVE_WAFER_AREA_MM2 * production_yield)
    else:
        raise ValueError(f"Unsupported device type for wafer equivalent calculation: {device_type}")
    return wafer_equivalent


def _get_year_index_from_series(series_map, production_year):
    years = [int(year) for year in series_map['years']]
    if production_year not in years:
        raise ValueError(f"Year must be between {years[0]} and {years[-1]}.")
    return years.index(production_year)


def calculate_ic_packaging_testing_impacts(production_year, n_ic=1, midpoint_scopes=None, spatial_awareness=False, location='Taiwan'):
    """
    Calculate backend IC packaging and testing impacts for a given production
    year and packaged/tested IC count.

    The SPIL data has both packaged-IC and tested-IC normalizations for the
    same annual totals. To avoid double counting, this helper uses the
    packaged-IC basis, which already represents the combined bumping,
    packaging, and testing backend workload per finished IC.
    """
    if n_ic < 0:
        raise ValueError("n_ic must be non-negative.")

    selected_midpoint_scopes = tuple(midpoint_scopes or MIDPOINT_SCOPES)
    invalid_scopes = [scope for scope in selected_midpoint_scopes if scope not in MIDPOINT_SCOPES]
    if invalid_scopes:
        raise ValueError(
            f"Unsupported midpoint scopes {invalid_scopes}. Available scopes: {MIDPOINT_SCOPES}"
        )

    year_idx = _get_year_index_from_series(spil_packaging_emissions, production_year)
    impacts = {midpoint: 0.0 for midpoint in selected_midpoint_scopes}
    location_specific_cfs = _get_location_specific_midpoint_cfs(location, spatial_awareness=spatial_awareness)

    electricity_consumption = (
        spil_packaging_emissions['total_electricity_consumption_kWh_per_packaged_ic'][year_idx] * n_ic
    )
    electricity_impacts = calculate_operational_impacts_given_energy(
        electricity_consumption,
        production_year,
        location=location,
        spatial_awareness=spatial_awareness,
    )
    for impact_type in selected_midpoint_scopes:
        impacts[impact_type] += electricity_impacts.get(impact_type, 0.0)

    # Treat water tons as m^3 because the reported annual water mass is effectively water volume.
    if 'WC' in impacts:
        impacts['WC'] += spil_packaging_emissions['total_water_consumption_ton_per_packaged_ic'][year_idx] * n_ic

    for metric_key, pollutant in SPIL_BACKEND_DIRECT_POLLUTANT_MAP.items():
        emission_mass = spil_packaging_emissions[metric_key][year_idx] * n_ic
        if emission_mass == 0:
            continue
        for impact_type in selected_midpoint_scopes:
            pollutant_cf = _get_midpoint_impact_factor(impact_type, pollutant, location_specific_cfs)
            if pollutant_cf is not None:
                impacts[impact_type] += pollutant_cf * emission_mass

    return impacts


def calculate_ic_packaging_testing_impact_breakdown(production_year, n_ic=1, midpoint_scopes=None, spatial_awareness=False, location='Taiwan'):
    """
    Return backend packaging/testing impacts separated into manufacturing
    scope 1 (direct onsite pollutants and water) and scope 2 (purchased
    electricity). Scope 3 is currently unresolved in this backend model.
    """
    if n_ic < 0:
        raise ValueError("n_ic must be non-negative.")

    selected_midpoint_scopes = tuple(midpoint_scopes or MIDPOINT_SCOPES)
    invalid_scopes = [scope for scope in selected_midpoint_scopes if scope not in MIDPOINT_SCOPES]
    if invalid_scopes:
        raise ValueError(
            f"Unsupported midpoint scopes {invalid_scopes}. Available scopes: {MIDPOINT_SCOPES}"
        )

    year_idx = _get_year_index_from_series(spil_packaging_emissions, production_year)
    breakdown = _empty_manufacturing_scope_breakdown(selected_midpoint_scopes)
    location_specific_cfs = _get_location_specific_midpoint_cfs(location, spatial_awareness=spatial_awareness)

    electricity_consumption = (
        spil_packaging_emissions['total_electricity_consumption_kWh_per_packaged_ic'][year_idx] * n_ic
    )
    electricity_impacts = calculate_operational_impacts_given_energy(
        electricity_consumption,
        production_year,
        location=location,
        spatial_awareness=spatial_awareness,
    )
    _add_midpoint_impacts(breakdown['scope2'], electricity_impacts, selected_midpoint_scopes)

    if 'WC' in breakdown['scope1']:
        breakdown['scope1']['WC'] += (
            spil_packaging_emissions['total_water_consumption_ton_per_packaged_ic'][year_idx] * n_ic
        )

    for metric_key, pollutant in SPIL_BACKEND_DIRECT_POLLUTANT_MAP.items():
        emission_mass = spil_packaging_emissions[metric_key][year_idx] * n_ic
        if emission_mass == 0:
            continue
        for impact_type in selected_midpoint_scopes:
            pollutant_cf = _get_midpoint_impact_factor(impact_type, pollutant, location_specific_cfs)
            if pollutant_cf is not None:
                breakdown['scope1'][impact_type] += pollutant_cf * emission_mass

    return breakdown


def calculate_cpu_manufacturing_impact_breakdown(
    cpu_specs,
    production_yield=0.875,
    spatial_awareness=False,
    calculate_upstream_materials=True,
    bom_template=None,
    assembly_transportation_distance=None,
    electricity_location='Taiwan',
    backend_location='Taiwan',
    **upstream_transportation_kwargs,
):
    """
    Calculate the CPU manufacturing impacts split into scope 1/2/3 buckets.
    
    Args:
        cpu_specs (dict): A dictionary containing CPU specifications. Expected keys are:
            - 'die_size_mm2': Die size in mm^2
            - 'technology_node_nm': Technology node in nm
            - 'production_year': Year of production
            - 'io_die_size_mm2': IO die size in mm^2 (optional, default is 0)
            - 'io_technology_node_nm': IO technology node in nm (optional, default is 0)
        
    Returns:
        dict: Scope 1/2/3 midpoint impacts for CPU manufacturing.
    """
    # Advanced nodes source: 5nm CMOS Production Technology Platform featuring full-fledged EUV - TSMC -IEDM 2019  
    # Mature node source: eBeam Initiative Annual Surveys (2015,2017,2020)

    year = cpu_specs['production_year']-2016  # Adjust year to match the index in the data
    if year < 0 or year > 8:
        raise ValueError("Year must be between 2016 and 2024.")

    breakdown = _empty_manufacturing_scope_breakdown(MIDPOINT_SCOPES)
    location_specific_cfs = _get_location_specific_midpoint_cfs('Taiwan', spatial_awareness=spatial_awareness)

    total_produce_units = (node_to_layer_masks_map[cpu_specs['technology_node_nm']] * (cpu_specs['die_size_mm2'] / EFFECTIVE_WAFER_AREA_MM2) + 
                           node_to_layer_masks_map[cpu_specs['io_die_technology_node_nm']] * (cpu_specs['io_die_size_mm2'] / EFFECTIVE_WAFER_AREA_MM2)) / production_yield

    # Acid gas emissions 
    for acid in tsmc_acid_emission_mix_ratio["2016"].keys():
        emission_mass = tsmc_acid_emission_mix_ratio[str(year+2016)][acid] * tsmc_emissions_macro['per_unit_acid_g/wafer-mask-layer'][year] * total_produce_units / 1000 # kg/device
        for impact_type in MIDPOINT_SCOPES:
            pollutant_cf = _get_midpoint_impact_factor(impact_type, acid, location_specific_cfs)
            if pollutant_cf is not None:
                breakdown['scope1'][impact_type] += pollutant_cf * emission_mass

    # SOx, NOx, VOC to air emissions
    for gas in ['SOx', 'NOx', 'VOC']:
        per_unit_emission = tsmc_emissions_macro[gas + '_mt'][year]*1e6 / (tsmc_emissions_macro['total_acid_mt'][year] * 1e6 / tsmc_emissions_macro['per_unit_acid_g/wafer-mask-layer'][year])
        emission_mass = per_unit_emission * total_produce_units / 1000 # kg/device
        for impact_type in MIDPOINT_SCOPES:
            pollutant_cf = _get_midpoint_impact_factor(impact_type, gas, location_specific_cfs)
            if pollutant_cf is not None:
                breakdown['scope1'][impact_type] += pollutant_cf * emission_mass
    
    # Wastewater discharge - Cu2+, NH4-N, and COD
    for pollutant in ['Cu2+', 'NH4-N', 'COD']:
        discharge_liter = tsmc_emissions_macro['per_unit_wastewater_L/wafer-mask-layer'][year] * total_produce_units # L/device
        discharge_mass = discharge_liter * tsmc_emissions_macro[pollutant + '_ppm'][year] * 1e-6 # kg/device
        for impact_type in MIDPOINT_SCOPES:
            pollutant_cf = _get_midpoint_impact_factor(impact_type, pollutant, location_specific_cfs)
            if pollutant_cf is not None:
                breakdown['scope1'][impact_type] += pollutant_cf * discharge_mass

    # Electricity consumption
    electricity_consumption = tsmc_electricity_consumption['unit_consumption_kWh/wafer-mask-layer'][year] * total_produce_units *(1- tsmc_electricity_consumption['renewable_energy_ratio_%'][year] * 1e-2) # Non-renewable kWh/device
    # Exclude CO2e from electricity consumption to avoid double counting scope 2 GWP. Include scope 2 water from electricity here
    electricity_impacts = calculate_operational_impacts_given_energy(
        electricity_consumption,
        year + 2016,
        location=electricity_location,
        spatial_awareness=spatial_awareness,
        excluded_pollutants=['CO2e'],
    )
    _add_midpoint_impacts(breakdown['scope2'], electricity_impacts)

    carbon_scope_split = _estimate_logic_fab_carbon_scope_split(cpu_specs, production_yield=production_yield)
    for scope_name in MANUFACTURING_SCOPE_KEYS:
        breakdown[scope_name]['GWP'] += carbon_scope_split[scope_name]

    direct_fab_water = cpu_specs['manufacturing_water_m3'] if 'manufacturing_water_m3' in cpu_specs else 0.0
    if direct_fab_water == 0.0 or direct_fab_water is None:
        direct_fab_water = cpu_specs.get('die_size_mm2', 0) * embodied_carbon_and_water['water_consumption_cmos_L_per_cm2_by_node'].get(str(cpu_specs['technology_node_nm']), 0) / production_yield / 100000 # L/cm^2 with mm^2 -> cm^2 and L -> m^3 conversion
    breakdown['scope1']['WC'] += direct_fab_water

    
    if calculate_upstream_materials:
        material_impacts = calculate_manufacturing_material_impacts({**cpu_specs, 'production_yield': production_yield}, spatial_awareness=spatial_awareness, **upstream_transportation_kwargs)
        material_impacts_total = _aggregate_midpoint_group_impacts(material_impacts)
        # Avoid calculating BoM twice for GPU
        bom_impacts = calculate_manufacturing_bom_impacts(cpu_specs, bom_template=bom_template, **upstream_transportation_kwargs) if cpu_specs.get('component_type', '') == 'CPU' else {}
        _add_midpoint_impacts(breakdown['scope3'], material_impacts_total)
        _add_midpoint_impacts(breakdown['scope3'], bom_impacts)

    backend_breakdown = calculate_ic_packaging_testing_impact_breakdown(
        cpu_specs['production_year'],
        n_ic=1,
        midpoint_scopes=MIDPOINT_SCOPES,
        spatial_awareness=spatial_awareness,
        location=backend_location,
    )
    for scope_name in MANUFACTURING_SCOPE_KEYS:
        _add_midpoint_impacts(breakdown[scope_name], backend_breakdown[scope_name])

    if assembly_transportation_distance is not None:
        assembly_transportation_impacts = calculate_transport_impact(
            cpu_specs.get('net_weight', 0),
            distance=assembly_transportation_distance,
        )
        _add_midpoint_impacts(breakdown['scope1'], assembly_transportation_impacts)

    return breakdown


def calculate_cpu_manufacturing_impacts(cpu_specs, production_yield=0.875, spatial_awareness=False, calculate_upstream_materials=True, bom_template=None, assembly_transportation_distance=None, **upstream_transportation_kwargs):
    return _sum_manufacturing_scope_breakdown(
        calculate_cpu_manufacturing_impact_breakdown(
            cpu_specs,
            production_yield=production_yield,
            spatial_awareness=spatial_awareness,
            calculate_upstream_materials=calculate_upstream_materials,
            bom_template=bom_template,
            assembly_transportation_distance=assembly_transportation_distance,
            **upstream_transportation_kwargs,
        )
    )


def debug_cpu_embodied_water_breakdown(cpu_specs, production_yield=0.875, spatial_awareness=False, debug_breakpoint=False, **upstream_transportation_kwargs):
    """
    Return a detailed embodied-water breakdown for a CPU and optionally pause at
    key checkpoints with ``breakpoint()`` for interactive debugging.
    """
    year = cpu_specs['production_year'] - 2016
    if year < 0 or year > 8:
        raise ValueError("Year must be between 2016 and 2024.")

    legacy_direct_fab_water_m3 = (
        cpu_specs.get('die_size_mm2', 0)
        * embodied_carbon_and_water['water_consumption_cmos_L_per_wafer_by_node'].get(str(cpu_specs['technology_node_nm']), 0)
        / production_yield
        / 1000
    )
    corrected_direct_fab_water_m3 = (
        cpu_specs.get('die_size_mm2', 0)
        * embodied_carbon_and_water['water_consumption_cmos_L_per_cm2_by_node'].get(str(cpu_specs['technology_node_nm']), 0)
        / production_yield
        / 100000
    )

    year_idx = cpu_specs['production_year'] - 2016
    total_produce_units = (
        node_to_layer_masks_map[cpu_specs['technology_node_nm']] * (cpu_specs['die_size_mm2'] / EFFECTIVE_WAFER_AREA_MM2)
        + node_to_layer_masks_map[cpu_specs['io_die_technology_node_nm']] * (cpu_specs['io_die_size_mm2'] / EFFECTIVE_WAFER_AREA_MM2)
    ) / production_yield
    electricity_consumption = (
        tsmc_electricity_consumption['unit_consumption_kWh/wafer-mask-layer'][year_idx]
        * total_produce_units
        * (1 - tsmc_electricity_consumption['renewable_energy_ratio_%'][year_idx] * 1e-2)
    )
    electricity_impacts = calculate_operational_impacts_given_energy(
        electricity_consumption,
        year + 2016,
        location='Taiwan',
        spatial_awareness=spatial_awareness,
        excluded_pollutants=['CO2e'],
    )
    material_impacts = calculate_manufacturing_material_impacts(
        {**cpu_specs, 'production_yield': production_yield},
        spatial_awareness=spatial_awareness,
        **upstream_transportation_kwargs,
    )
    material_impacts_total = _aggregate_midpoint_group_impacts(material_impacts)
    backend_impacts = calculate_ic_packaging_testing_impacts(
        cpu_specs['production_year'],
        n_ic=1,
        midpoint_scopes=MIDPOINT_SCOPES,
        spatial_awareness=spatial_awareness,
    )
    bom_impacts = calculate_manufacturing_bom_impacts(cpu_specs) if cpu_specs.get('component_type', '') == 'CPU' else {}

    if debug_breakpoint:
        breakpoint()  # Inspect legacy per-wafer usage versus corrected per-cm^2 water before aggregation.

    breakdown = {
        'device_name': cpu_specs.get('name'),
        'production_yield': production_yield,
        'legacy_direct_fab_water_m3': legacy_direct_fab_water_m3,
        'corrected_direct_fab_water_m3': corrected_direct_fab_water_m3,
        'tsmc_scope2_electricity_wc_m3': electricity_impacts.get('WC', 0.0),
        'upstream_silicon_wc_m3': material_impacts.get('silicon', {}).get('WC', 0.0),
        'upstream_chemicals_wc_m3': material_impacts.get('chemicals', {}).get('WC', 0.0),
        'upstream_materials_wc_m3_total': material_impacts_total.get('WC', 0.0),
        'backend_packaging_testing_wc_m3': backend_impacts.get('WC', 0.0),
        'bom_wc_m3': bom_impacts.get('WC', 0.0),
    }
    breakdown['total_embodied_wc_m3'] = (
        breakdown['corrected_direct_fab_water_m3']
        + breakdown['tsmc_scope2_electricity_wc_m3']
        + breakdown['upstream_materials_wc_m3_total']
        + breakdown['backend_packaging_testing_wc_m3']
        + breakdown['bom_wc_m3']
    )

    if debug_breakpoint:
        breakpoint()  # Inspect the assembled embodied-water breakdown before returning it.

    return breakdown

def calculate_cpu_manufacturing_pollutants(cpu_specs, production_yield=0.875):
    """
    Calculate the CPU manufacturing pollutants based on the provided specifications. 
        - Gas emissions
        - Wastewater discharge
    
    Args:
        cpu_specs (dict): A dictionary containing CPU specifications. Expected keys are:
            - 'die_size_mm2': Die size in mm^2
            - 'technology_node_nm': Technology node in nm
            - 'production_year': Year of production
            - 'io_die_size_mm2': IO die size in mm^2 (optional, default is 0)
            - 'io_technology_node_nm': IO technology node in nm (optional, default is 0)
        
    Returns:
        pollutants (dict): A dictionary containing the calculated manufacturing pollutants.
    """

    year = cpu_specs['production_year']-2016  # Adjust year to match the index in the data
    if year < 0 or year > 8:
        raise ValueError("Year must be between 2016 and 2024.")

    pollutants = {}

    total_produce_units = (node_to_layer_masks_map[cpu_specs['technology_node_nm']] * (cpu_specs['die_size_mm2'] / EFFECTIVE_WAFER_AREA_MM2) + 
                           node_to_layer_masks_map[cpu_specs['io_die_technology_node_nm']] * (cpu_specs['io_die_size_mm2'] / EFFECTIVE_WAFER_AREA_MM2)) / production_yield

    # Acid gas emissions 
    for acid in tsmc_acid_emission_mix_ratio["2016"].keys():
        if acid != 'hydrocarbon':  # Avoid double counting VOCs
            emission_mass = tsmc_acid_emission_mix_ratio[str(year+2016)][acid] * tsmc_emissions_macro['per_unit_acid_g/wafer-mask-layer'][year] * total_produce_units / 1000 # kg/device
            pollutants[acid] = emission_mass

    # SOx, NOx, VOC air emissions
    for gas in ['SOx', 'NOx', 'VOC']:
        per_unit_emission = tsmc_emissions_macro[gas + '_mt'][year]*1e6 / (tsmc_emissions_macro['total_acid_mt'][year] * 1e6 / tsmc_emissions_macro['per_unit_acid_g/wafer-mask-layer'][year])
        emission_mass = per_unit_emission * total_produce_units / 1000 # kg/device
        pollutants[gas] = emission_mass
    
    # Wastewater discharge - Cu2+, NH4-N, and COD
    for pollutant in ['Cu2+', 'NH4-N', 'COD']:
        discharge_liter = tsmc_emissions_macro['per_unit_wastewater_L/wafer-mask-layer'][year] * total_produce_units # L/device
        discharge_mass = discharge_liter * tsmc_emissions_macro[pollutant + '_ppm'][year] * 1e-6 # kg/device
        pollutants[pollutant] = discharge_mass

    return pollutants

def calculate_storage_manufacturing_impact_breakdown(
    storage_type,
    production_year,
    capacity,
    production_yield=0.875,
    unit='GB',
    HBM_edition=None,
    spatial_awareness=False,
    calculate_upstream_materials=True,
    electricity_location='Korea',
    **upstream_transportation_kwargs,
):
    """
    Calculate storage manufacturing impacts split into scope 1/2/3 buckets.
    
    Args:
        storage_type (str): Type of storage ('DRAM', 'NAND', 'SSD', 'HDD').
        production_year (int): Year of production.
        capacity (float): Storage capacity in GB.
        production_yield (float): Production yield for the storage device.
        unit (str): Unit of capacity ('GB' or 'TB').
        HBM_edition (str): Edition of HBM storage (if applicable).
        calculate_upstream_materials (bool): Whether to calculate upstream materials impacts.
        **upstream_transportation_kwargs: Additional keyword arguments for upstream transportation calculations.


    Returns:
        dict: Scope 1/2/3 midpoint impacts for storage manufacturing.
    """
    if storage_type not in ['DRAM', 'SSD', 'HBM', 'HDD']:
        raise ValueError("Invalid storage type. Must be one of ['DRAM', 'SSD', 'HBM', 'HDD'].")
    
    if unit not in ['GB', 'TB']:
        raise ValueError("Invalid unit. Must be either 'GB' or 'TB'.")
    if unit == 'TB':
        capacity *= 1000
    
    if storage_type == 'HDD':
        # Validate HDD year against available by-stage data loaded from JSON.
        _get_hdd_stage_impact_factors_for_year(production_year)
    elif production_year < 2016 or production_year > 2024:
        raise ValueError("Year must be between 2016 and 2024 for DRAM/SSD/HBM.")
    
    if capacity <= 0:
        raise ValueError("Capacity must be a positive number.")
    
    # Initialize impacts
    breakdown = _empty_manufacturing_scope_breakdown(MIDPOINT_SCOPES)
    if storage_type in ['DRAM', 'HBM', 'SSD']:
        location_specific_cfs = _get_location_specific_midpoint_cfs('Korea', spatial_awareness=spatial_awareness)
        equivalent_wafers = get_equivalent_wafers({'device_type': storage_type, 'capacity': capacity, 'production_year': production_year, 'hbm_type': HBM_edition}, production_yield=production_yield)
        
        # Get the year index
        year_idx = production_year - 2016
        device_revenue_ratio = hynix_production_data['dram_revenue_ratio'][year_idx] if storage_type in ['DRAM', 'HBM'] else hynix_production_data['nand_revenue_ratio'][year_idx]
        device_wafer_capacity = hynix_production_data['estimated_dram_k_wafers'][year_idx] * 1000 if storage_type in ['DRAM', 'HBM'] else hynix_production_data['nand_k_wafers'][year_idx] * 1000

        for pollutant in dram_emissions[production_year].keys():
            pollutant_mass = dram_emissions[production_year][pollutant] * equivalent_wafers # kg/device
            for impact_type in MIDPOINT_SCOPES:
                pollutant_cf = _get_midpoint_impact_factor(impact_type, pollutant, location_specific_cfs)
                if pollutant_cf is not None:
                    breakdown['scope1'][impact_type] += pollutant_cf * pollutant_mass

        wafer_electricity_consumption = hynix_production_data['electricity_consumption_GWh'][year_idx] * 1e6 * device_revenue_ratio / device_wafer_capacity * (1-hynix_production_data["renewable_energy_ratio_%"][year_idx]*0.01)# kWh/wafer non-renewable
        electricity_consumption = wafer_electricity_consumption * equivalent_wafers # kWh/device
        electricity_impacts = calculate_operational_impacts_given_energy(
            electricity_consumption,
            year_idx + 2016,
            location=electricity_location,
            spatial_awareness=spatial_awareness,
            excluded_pollutants=['CO2e'],
        )
        _add_midpoint_impacts(breakdown['scope2'], electricity_impacts)

        fab_carbon_split = _estimate_storage_fab_carbon_scope_split(
            storage_type,
            production_year,
            equivalent_wafers,
            HBM_edition=HBM_edition,
        )
        breakdown['scope1']['GWP'] += fab_carbon_split['scope1']
        breakdown['scope2']['GWP'] += fab_carbon_split['scope2']

        # Water consumption Scope 1
        water_per_wafer = hynix_production_data['total_water_consumed_1000_m3'][year_idx] * 1000 * device_revenue_ratio / device_wafer_capacity # m3/wafer
        breakdown['scope1']['WC'] += water_per_wafer * equivalent_wafers # m3/device

        if calculate_upstream_materials:
            material_impacts = calculate_manufacturing_material_impacts({'device_type': storage_type, 'capacity': capacity, 'production_year': production_year, 'hbm_type': HBM_edition, 'production_yield': production_yield}, spatial_awareness=spatial_awareness, **upstream_transportation_kwargs)
            material_impacts_total = _aggregate_midpoint_group_impacts(material_impacts)
            _add_midpoint_impacts(breakdown['scope3'], material_impacts_total)

    else:
        hdd_stage_impact_factors = _get_hdd_stage_impact_factors_for_year(production_year)
        for impact_type in MIDPOINT_SCOPES:
            if impact_type in hdd_stage_impact_factors:
                breakdown['scope1'][impact_type] += hdd_stage_impact_factors[impact_type].get('manufacturing', 0.0) * capacity

    return breakdown


def calculate_storage_manufacturing_impacts(storage_type, production_year, capacity, production_yield=0.875, unit='GB', HBM_edition=None, spatial_awareness=False, calculate_upstream_materials=True, **upstream_transportation_kwargs):
    return _sum_manufacturing_scope_breakdown(
        calculate_storage_manufacturing_impact_breakdown(
            storage_type,
            production_year,
            capacity,
            production_yield=production_yield,
            unit=unit,
            HBM_edition=HBM_edition,
            spatial_awareness=spatial_awareness,
            calculate_upstream_materials=calculate_upstream_materials,
            **upstream_transportation_kwargs,
        )
    )

def calculate_storage_manufacturing_pollutants(storage_type, production_year, capacity, production_yield=0.875, unit='GB', HBM_edition=None):
    """
    Calculate the manufacturing pollutants (direct fab emissions) for different storage types based on the provided specifications.
    
    Args:
        storage_type (str): Type of storage ('DRAM', 'NAND', 'SSD').
        production_year (int): Year of production.
        capacity (float): Storage capacity in GB.
        
    Returns:
        pollutants (dict): A dictionary containing the calculated manufacturing pollutants.
    """
    if storage_type not in ['DRAM', 'SSD', 'HBM']:
        raise ValueError("Invalid storage type. Must be one of ['DRAM', 'SSD', 'HBM'].")
    
    if unit not in ['GB', 'TB']:
        raise ValueError("Invalid unit. Must be either 'GB' or 'TB'.")
    if unit == 'TB':
        capacity *= 1000
    
    if production_year < 2016 or production_year > 2024:
        raise ValueError("Year must be between 2016 and 2024.")
    
    if capacity <= 0:
        raise ValueError("Capacity must be a positive number.")
    
    # Initialize pollutants
    pollutants = {}
    
    # Get the year index
    year_idx = production_year - 2016

    if storage_type in ['DRAM', 'HBM', 'SSD']:
        equivalent_wafers = get_equivalent_wafers({'device_type': storage_type, 'capacity': capacity, 'production_year': production_year, 'hbm_type': HBM_edition}, production_yield=production_yield)
        device_emissions = dram_emissions if storage_type in ['DRAM', 'HBM'] else nand_emissions
        for pollutant in device_emissions[production_year].keys():
            pollutant_mass = device_emissions[production_year][pollutant] * equivalent_wafers # kg/device
            pollutants[pollutant] = pollutant_mass

    return pollutants

def calculate_gpu_manufacturing_impact_breakdown(gpu_specs, memory_production_yield=0.875, spatial_awareness=False, calculate_upstream_materials=True, bom_template=None, assembly_transportation_distance=None, **upstream_transportation_kwargs):
    """
    Calculate the GPU manufacturing impacts split into scope 1/2/3 buckets.
    
    GPU is basically a special CPU die plus HBM. We will use the existing functions to calculate the overall impact.
    Args:
        gpu_specs (dict): A dictionary containing GPU specifications. Superset of CPU specs.
            - 'die_size_mm2': Die size in mm^2
            - 'technology_node_nm': Technology node in nm
            - 'production_year': Year of production
            - 'die_production_yield': Production yield for the die. Could be very low for A100 / H100.
            - 'hbm_capacity_GB': HBM capacity in GB 
            - 'hbm_type': HBM type (e.g., 'HBM2', 'HBM3')
    """
    die_breakdown = calculate_cpu_manufacturing_impact_breakdown(
        gpu_specs,
        production_yield=gpu_specs['die_production_yield'],
        spatial_awareness=spatial_awareness,
        calculate_upstream_materials=calculate_upstream_materials,
        assembly_transportation_distance=None,
        **upstream_transportation_kwargs,
    )
    hbm_breakdown = calculate_storage_manufacturing_impact_breakdown(
        'HBM',
        gpu_specs['production_year'],
        gpu_specs['hbm_capacity_GB'],
        production_yield=memory_production_yield,
        HBM_edition=gpu_specs['hbm_type'],
        spatial_awareness=spatial_awareness,
        calculate_upstream_materials=calculate_upstream_materials,
        **upstream_transportation_kwargs,
    )
    breakdown = _empty_manufacturing_scope_breakdown(MIDPOINT_SCOPES)
    for scope_name in MANUFACTURING_SCOPE_KEYS:
        _add_midpoint_impacts(breakdown[scope_name], die_breakdown[scope_name])
        _add_midpoint_impacts(breakdown[scope_name], hbm_breakdown[scope_name])

    if calculate_upstream_materials:
        bom_impacts = calculate_manufacturing_bom_impacts(gpu_specs, bom_template=bom_template, **upstream_transportation_kwargs)
        _add_midpoint_impacts(breakdown['scope3'], bom_impacts)
    
    if assembly_transportation_distance is not None:
        transportation_impacts = calculate_transport_impact(gpu_specs.get('net_weight', 0), distance=assembly_transportation_distance)
        _add_midpoint_impacts(breakdown['scope1'], transportation_impacts)

    return breakdown


def calculate_gpu_manufacturing_impacts(gpu_specs, memory_production_yield=0.875, spatial_awareness=False, calculate_upstream_materials=True, bom_template=None, assembly_transportation_distance=None, **upstream_transportation_kwargs):
    return _sum_manufacturing_scope_breakdown(
        calculate_gpu_manufacturing_impact_breakdown(
            gpu_specs,
            memory_production_yield=memory_production_yield,
            spatial_awareness=spatial_awareness,
            calculate_upstream_materials=calculate_upstream_materials,
            bom_template=bom_template,
            assembly_transportation_distance=assembly_transportation_distance,
            **upstream_transportation_kwargs,
        )
    )

def calculate_gpu_manufacturing_pollutants(gpu_specs, memory_production_yield=0.875):
    """
    Calculate the GPU manufacturing pollutants based on the provided specifications. 
    
    GPU is basically a special CPU die plus HBM. We will use the existing functions to calculate the overall pollutants.
    Args:
        gpu_specs (dict): A dictionary containing GPU specifications. Superset of CPU specs.
            - 'die_size_mm2': Die size in mm^2
            - 'technology_node_nm': Technology node in nm
            - 'production_year': Year of production
            - 'die_production_yield': Production yield for the die. Could be very low for A100 / H100.
            - 'hbm_capacity_GB': HBM capacity in GB
            - 'hbm_type': HBM type (e.g., 'HBM2', 'HBM3')
    """
    die_pollutants = calculate_cpu_manufacturing_pollutants(gpu_specs, production_yield=gpu_specs['die_production_yield'])
    hbm_pollutants = calculate_storage_manufacturing_pollutants('HBM', gpu_specs['production_year'], gpu_specs['hbm_capacity_GB'], production_yield=memory_production_yield, HBM_edition=gpu_specs['hbm_type'])
    pollutants = {}
    for key in set(die_pollutants.keys()).union(set(hbm_pollutants.keys())):
        pollutants[key] = die_pollutants.get(key, 0) + hbm_pollutants.get(key, 0)
    return pollutants


def calculate_manufacturing_impact_breakdown(device_specs, spatial_awareness=False, calculate_upstream_materials=True, bom_template=None, **manufacturing_kwargs):
    """
    Return scope 1/2/3 manufacturing midpoint impacts for a device without
    changing the legacy aggregated manufacturing totals.
    """
    component_type = device_specs['component_type']
    if component_type == 'CPU':
        return calculate_cpu_manufacturing_impact_breakdown(
            device_specs,
            production_yield=device_specs.get('production_yield', 0.875),
            spatial_awareness=spatial_awareness,
            calculate_upstream_materials=calculate_upstream_materials,
            bom_template=bom_template,
            **manufacturing_kwargs,
        )
    if component_type == 'GPU':
        return calculate_gpu_manufacturing_impact_breakdown(
            device_specs,
            memory_production_yield=device_specs.get('memory_production_yield', 0.875),
            spatial_awareness=spatial_awareness,
            calculate_upstream_materials=calculate_upstream_materials,
            bom_template=bom_template,
            **manufacturing_kwargs,
        )
    if component_type in ['SSD', 'HDD', 'DRAM']:
        return calculate_storage_manufacturing_impact_breakdown(
            device_specs['component_type'],
            device_specs['production_year'],
            device_specs['capacity'],
            production_yield=device_specs.get('production_yield', 0.875),
            spatial_awareness=spatial_awareness,
            calculate_upstream_materials=calculate_upstream_materials,
            **manufacturing_kwargs,
        )
    raise ValueError("Invalid component type. Must be one of ['CPU', 'GPU', 'SSD', 'HDD', 'DRAM'].")

distance = {
    'default': {
        # East Asia manufacture, US usage & recycling, air transport for high-value components
        # Estimated based on Dell R740 LCA report
        'Truck': 1300,  # km, 100 for fab-to-shipyard, 1200 for port-to-DC
        'Ship': 0,  # km
        'Air': 14000,    # km, East Asia to US west coast
    },
}
LOCATION_SPECIFIC_WATER_ENDPOINT_ALIASES = {
    'US': ('United States', 'USA', 'US'),
    'CA': ('United States',),
    'TX': ('United States',),
    'VA': ('United States',),
    'IA': ('United States',),
    'MIDWEST': ('United States',),
    'SPAIN': ('Spain',),
    'GERMANY': ('Germany',),
    'FRANCE': ('France',),
    'CHINA': ('China',),
    'TAIWAN': ('Taiwan', 'China'),
    'KOREA': ('Korea, Republic of', 'South Korea', 'Korea'),
    'JAPAN': ('Japan',),
    'SWEDEN': ('Sweden',),
}

def _copy_distance_map(distance_map):
    return {mode: distance_map.get(mode, 0) for mode in TRANSPORT_DISTANCE_KEYS}


def _normalize_transport_regions(use_region=None, eol_region=None, transport_region=None):
    if transport_region is not None:
        if use_region is None:
            use_region = transport_region
        if eol_region is None:
            eol_region = transport_region
    return use_region, eol_region


def get_device_transport_family(device_specs):
    component_type = device_specs.get('component_type', device_specs.get('device_type'))
    if component_type in LOGIC_COMPONENT_TYPES:
        return 'logic'
    if component_type in MEMORY_STORAGE_COMPONENT_TYPES:
        return 'memory_storage'
    raise ValueError(f"Unsupported component type for transport routing: {component_type!r}")


def _get_transport_region_route(region_name):
    routes = regional_device_transport_routes['regions']
    if region_name not in routes:
        raise ValueError(
            f"Unsupported transport region {region_name!r}. "
            f"Available regions: {sorted(routes.keys())}"
        )
    return routes[region_name]


def _build_regional_manufacture_to_use_distance(transport_family, region_name):
    route = _get_transport_region_route(region_name)
    fixed_origin_leg = regional_device_transport_routes['metadata']['fixed_origin_legs'][transport_family]
    air_distance_key = 'air_distance_from_taipei_km' if transport_family == 'logic' else 'air_distance_from_seoul_km'
    return {
        'Truck': fixed_origin_leg['truck_km'] + route['last_mile_truck_km'],
        'Ship': 0,
        'Air': route[air_distance_key],
    }


def resolve_transport_distances(device_specs, use_region=None, eol_region=None, transport_region=None):
    """
    Resolve route distances for manufacture-to-use and use-to-EoL stages.

    If no regional overrides are provided, this preserves the existing device-spec
    behavior. ``transport_region`` is a convenience alias that fills both
    ``use_region`` and ``eol_region`` unless either one is provided explicitly.
    """
    use_region, eol_region = _normalize_transport_regions(
        use_region=use_region,
        eol_region=eol_region,
        transport_region=transport_region,
    )
    transport_family = get_device_transport_family(device_specs)
    resolved = {
        'transport_family': transport_family,
        'use_region': use_region,
        'eol_region': eol_region,
        'manufacture_to_use': _copy_distance_map(device_specs['to_use_distance']) if use_region is None else None,
        'use_to_eol': _copy_distance_map(device_specs.get('to_recycle_distance', distance['default'])) if eol_region is None else None,
    }

    if use_region is not None:
        route = _get_transport_region_route(use_region)
        resolved['manufacture_to_use'] = _build_regional_manufacture_to_use_distance(transport_family, use_region)
        resolved['use_region_class'] = route['region_class']
        resolved['use_destination_air_hub'] = route['destination_air_hub']

    if eol_region is not None:
        route = _get_transport_region_route(eol_region)
        eol_truck_km = route.get('eol_truck_km')
        if eol_truck_km is None:
            raise ValueError(
                f"No use-to-EoL truck distance is defined for region {eol_region!r}. "
                "The current route dataset only assigns EoL defaults for US and EU regions. "
                "Pass eol_region=None to keep the device-spec fallback or specify a supported EoL region."
            )
        resolved['use_to_eol'] = {'Truck': eol_truck_km, 'Ship': 0, 'Air': 0}
        resolved['eol_region_class'] = route['region_class']

    return resolved


def calculate_transport_impact(mass, distance=distance['default'], mass_unit='g', midpoint_scopes=None, warn_missing=False):
    """
    Calculate transportation impact given mass and distance
    Source: ELCD 3.2 Transportation
    
    Parameters:
    mass: float - mass of the component
    distance: dict - distances for different transport modes
    mass_unit: str - unit of mass ('g', 'kg', 't', default 'g')
    year: int - year for which to calculate impact (2016-2024)
    
    Returns:
    dict - impact values for each transport mode and impact type
    """
    # Convert mass to kg
    conversion = {
        'g': 1e-3,
        'kg': 1,
        't': 1e3
    }
    
    if mass_unit not in conversion:
        raise ValueError("Mass unit must be 'g', 'kg', or 't'")
    
    # mass_in_tonnes = mass * conversion[mass_unit]
    mass_in_kg = mass * conversion[mass_unit]  # convert to kg
    

    selected_midpoint_scopes = tuple(midpoint_scopes or MIDPOINT_SCOPES)
    impacts = {midpoint: 0.0 for midpoint in selected_midpoint_scopes}
    for transport_means, dist in distance.items():
        if transport_means not in cf_transportation:
            if warn_missing:
                print(f"Warning: Missing transport mode {transport_means}, skipping calculation.")
            continue

        transport_factors = cf_transportation[transport_means]
        reference_weight = transport_factors.get('reference_weight_kg', 1) or 1
        reference_distance = transport_factors.get('reference_distance_km', 1) or 1

        for impact_type in selected_midpoint_scopes:
            midpoint_factor = transport_factors.get(impact_type)
            if midpoint_factor is None:
                if warn_missing:
                    print(f"Warning: Missing CF for {transport_means} and {impact_type}, skipping calculation.")
                continue
            impacts[impact_type] += midpoint_factor * mass_in_kg / reference_weight * dist / reference_distance

    return impacts

    
def calculate_manufacturing_pollutant_impacts(device_specs):
    """
    Calculate the manufacturing pollutant impacts of a device based on its specifications.
    
    Args:
        device_specs (dict): A dictionary containing device specifications. Expected keys are:
            - 'component_type': Type of component ('CPU', 'GPU', 'SSD', 'HDD', 'DRAM')
            - 'net_weight': Mass of the component in grams 
            - 'production_year': Year of production (2016-2024)
            - Other keys required for specific component types
            
    Returns:
        dict: A dictionary containing the calculated manufacturing pollutant midpoint impacts.
    """
    component_type = device_specs['component_type']
    
    if component_type == 'CPU':
        pollutants = calculate_cpu_manufacturing_pollutants(device_specs)
    elif component_type == 'GPU':
        pollutants = calculate_gpu_manufacturing_pollutants(device_specs)
    elif component_type in ['SSD', 'HDD', 'DRAM']:
        pollutants = calculate_storage_manufacturing_pollutants(device_specs)

    else:
        raise ValueError("Unsupported component type for pollutant calculation.")

    impacts = {k:0.0 for k in MIDPOINT_SCOPES}
    for pollutant, mass in pollutants.items():
        for impact_type in MIDPOINT_SCOPES:
            pollutant_cf = _get_midpoint_impact_factor(impact_type, pollutant)
            if pollutant_cf is not None:
                impacts[impact_type] += pollutant_cf * mass
    return impacts

def calculate_manufacturing_material_impacts(device_specs, bit_density=None, silicon_wafer_region='Japan', chemical_region='Japan', silicon_wafer_transportation=None, chemical_transportation=None, spatial_awareness=False):
    """
    Calculate the material procurement impacts (silicon and process chemicals) of a device based on its specifications.
    This function only models the embodied electricity impact for silicon wafer and process chemicals procurement. 
    The supply chain transportation impacts can be optionally calculated.
    The typical silicon wafer process is: quartz (SiO2) + carbon (C) -> metallurgical grade silicon + Chlorine (Cl2) -> HSiCl3 + Hydrogen (H2) -> polysilicon (+HCl) -> single-crystal silicon -> silicon wafer. The byproducts are either harmless or can be collected and sold to other industries, which should incur minimal environmental impact. The main input chemicals, Cl2 and H2 in the supply chain, are usually produced via electrolysis -- which also do not release harzoudous substances directly but only have indirect impacts from electricity generation. Therefore, the main impact comes from the electricity use in the silicon wafer manufacturing process, which is modeled based on the reported energy consumption per unit area of wafer for different technology nodes and years.
    The process chemicals include photoresist, developer, etching gases, cleaning chemicals, etc. 
    The extraction of resource in ground, and construction of infratructure, manufacturing of equipment, and other capital goods are not included for modeling the cradle-to-gate impact of silicon wafer and process chemicals procurement, due to the lack of data and the relatively smaller contribution compared to the energy use impact in the semiconductor manufacturing stage. 
    If background CFs are used (e.g., Ecoinvent/Sphera), the impact will be higher. 

    For the silicon wafer embodied electricity, the calculation is based on 'The 1.7 Kilogram Microchip: Energy and Material Use in the Production of Semiconductor Devices', ES&T 2002 (for mature nodes above 28nm), and 'Cradle-togate Life Cycle Assessment of CMOS Logic Technologies', IEDM 2023 (for advanced nodes 28nm and below). 
    For process chemicals, the mature nodes use 'Life-cycle Assessment of Semiconductor', Boyd 2012, and the advanced nodes use the same IEDM 2023 paper by intersecting the non-wafer scope-3 energy consumption between 28nm and 3nm nodes using # of process steps as the independent variable, and then apply the same energy-based approach to estimate the chemical procurement impact for advanced nodes.
    
    Args:
        device_specs (dict): A dictionary containing device specifications. Expected keys are:
            - 'component_type': Type of component ('CPU', 'GPU', 'SSD', 'HDD', 'DRAM', 'HBM'). Note that to compute the impact for GPU, the function should be called twice for the compute die and the HBM separately 
            - 'production_year': Year of production (2016-2024)
            - 'bit_density': Bit density of the component (optional, used for equivalent wafer calculation when the storage node is not in the default 2016-2024 range)
            - Other keys required for specific component types
        silicon_wafer_region (str): The region where the whole silicon wafer supply chain is assumed to be located.
        chemical_region (str): The region where the process chemicals are assumed to be produced.
        silicon_wafer_transportation (dict, optional): Transportation modes and distances for silicon wafer supply chain.
        chemical_transportation (dict, optional): Transportation modes and distances for process chemical supply chain.

    Returns:
        dict: A dictionary containing the calculated material impacts.
    """
    component_type = device_specs.get('component_type', device_specs.get('device_type'))
    if component_type is None:
        raise KeyError("device_specs must include either 'component_type' or 'device_type'.")
    assert component_type in ['CPU', 'GPU', 'SSD', 'DRAM', 'HBM'], "Unsupported component type for material impact calculation."

    impacts = {'silicon': None, 'chemicals': None}
    EMBODIED_ENERGY_PER_SILICON_WAFER_AREA = {
        'mature': 0.34, # kWh per cm^2,
        'advanced': 0.426  
    }
    KG_SILICON_PER_CM2 = 0.34 / 2130 # convert kWh/cm^2 to kg silicon/cm^2 using the energy intensity of silicon wafer production from ES&T 2002 paper. This is used to estimate the transportation impact based on mass.
    KG_CHEMICALS_PER_WAFER_LOGIC_45NM =  (3.46e-2+3.3e-3+1.01+1.76e-1+2.05e-4+4.09e-4+7.5e-4+1.59e-1+2.35e-3+2.84e-4+6.07e-4+7.1e-5+3.09e-2+1e-2+2.88e-1+5.9e-3+1.68e-2+1.02e-1+3.27e-3+7.4e-3+4.31+1.92+1.66+23.6+23.3+7.46e-1+7.46e-1+3.07e-2+5.45e-1+3.38e-2+2.62+1.53e-3+1.07e-7+3.21e-1+1.22e-2+7.58e-5+1.31e-2+4.85e-2+1.15e-3+3e-3+2.62e-2+3.35e-2+3.72e-2+2.02e-2+2.82e-1+3.16e-2+3.44e-1+3.27e-1+3.52e-2+5.35e-5+1.2e-2+2.36e-2+1.15e-1+253+5.35+6.97+1.11+5.03+2.26e-4+4.38e-3+1.2e-4+1.01e-2+2.35e-3+9.96e-4+4.81e-7+7.16e-3+3.41e-3+2.43e-1) / 1000 # kg/wafer, estimated based on the chemical use data in Table B.17 of Boyd book for 45nm logic node, which is considered as the representative mature node. This is used to estimate the transportation impact based on mass.
    KG_CHEMICALS_PER_WAFER_DRAM_57NM = (2.19e-2+3.41e-3+6.48e-1+2.05e-1+2.56e-4+1.31e-3+3.77e-3+2.49e-3+6.21e-4+7.71e-2+1.75e-2+2.88e-1+1.85e-2+2.13e-1+7.04e-1+1.46+5.37e-1+13.1+2.18e-1+1.21e-1+10.5+1e-3+1.07e-7+8.48e-4+9.23e-4+1.31e-2+1.15e-3+2.08e-3+1.32e-3+3.35e-2+3.72e-2+2.02e-2+2.82e-1+3.16e-2+3.44e-1+3.27e-1+3.52e-2+1.07e-4+1.3e-1+99.2+4.64+6.99+15.1+2.89e-4+5.6e-3+1.54e-4+1.29e-2+2.35e-3+9.83e-4+4.81e-7+2.83e-4+2.43e-1) / 1000 # kg/wafer, estimated based on the chemical use data in Table D.12 of Boyd book for 57nm DRAM node, which is considered as the representative mature node for memory devices.

    required_wafer_area_cm2 = get_equivalent_wafers(device_specs, bit_density=bit_density) * EFFECTIVE_WAFER_AREA_MM2 / 100 # convert mm^2 to cm^2
    if component_type in ['CPU', 'GPU']:
        logic_node_nm = float(device_specs['technology_node_nm'])
        energy_per_cm2 = EMBODIED_ENERGY_PER_SILICON_WAFER_AREA['advanced'] if logic_node_nm <= 28 else EMBODIED_ENERGY_PER_SILICON_WAFER_AREA['mature']
    else:
        energy_per_cm2 = EMBODIED_ENERGY_PER_SILICON_WAFER_AREA['mature']  if bit_density is not None and bit_density < dram_bit_density_by_year[0] else EMBODIED_ENERGY_PER_SILICON_WAFER_AREA['advanced']
    silicon_wafer_energy = energy_per_cm2 * required_wafer_area_cm2 # kWh

    # Logic device have scope 3 GWP in calculate_cpu_manufacturing_impacts() that already accounts for the embodied carbon from silicon wafer production electricity use, so we only calculate other impact types for silicon wafer procurement for CPU and GPU to avoid double counting. For memory devices without scope 3 GWP in calculate_storage_manufacturing_impacts(), we calculate all impact types for silicon wafer procurement.
    impacts['silicon'] = calculate_operational_impacts_given_energy(silicon_wafer_energy, device_specs['production_year'], silicon_wafer_region, spatial_awareness=spatial_awareness,excluded_pollutants=['CO2e'] if component_type in ['CPU', 'GPU'] else []) 
    if silicon_wafer_transportation is not None:
        transportation_impacts = calculate_transport_impact(required_wafer_area_cm2 * KG_SILICON_PER_CM2, distance=silicon_wafer_transportation) # convert cm^2 to kg using the estimated conversion factor
        for impact_type in MIDPOINT_SCOPES:
            impacts['silicon'][impact_type] += transportation_impacts.get(impact_type, 0.0)

    # Process chemicals
    PER_UNIT_PROCESS_CHEMICAL_ENERGY = {
        'logic_mature_lower_bound':  4.2 / 10.7 * 443 * 0.88, # MJ/die attributed to process chemicals for 45nm; MJ/kWh conversion; Good die per wafer ("net yield" in Boyd book Table 3.14); finished wafer/wafer start ("line yield" in Boyd book Table 3.14); final result in kWh/wafer
        'logic_mature_upper_bound': 1e-4 * 10 ** (201/222) * 443*1000/4.3654/0.45359237 * 0.88, # kg NOx/die attributed to process chemicals for 45nm estimated based on the pixel heights in Figure 4.4 of Boyd book,  x1000 to convert kg to g, 4.3654 is Pacific Gas and Electric's fossil fuel output-based NOx emission factor in lb/MWh, 0.45359237 is the conversion factor from lb/MWh to kg/MWh, and 0.88 is the same correction factor for line yield as the lower bound
        'logic_advanced_28nm': 1.38*(1-0.288-0.088-0.141-0.14-0.076-0.025)/0.454, # kWh/cm^2 for 28nm node estimated based on the non-wafer scope-3 energy consumption in IEDM 2023 paper Figure 7
        'logic_advanced_3nm': 2.85*(1-0.38-0.111-0.163-0.066-0.032-0.068)/0.454, # kWh/cm^2 for 3nm node estimated based on the non-wafer scope-3 energy consumption in IEDM 2023 paper Figure 7
        'dram': 0.1*10**(242/266)/12 /4.096 * 0.21 * 100, # 0.1*10**(242/266)/12 is the embodied kWh per 512MB DRAM attributed to process chemicals for 57 nm based on the pixel heights in Figure 6.1 of Boyd book, divided by 4.096 to convert to kWh/Gb, and multiplied by 0.21 Gb/mm^2 to get kWh/mm^2, multiplied by 100 to get kWh/cm^2 (though nonsense bit density data, but gives closer gap of 57nm at 0.34 kWh/cm^2 compared to 28nm logic ~0.74 kWh/cm^2, and also keeps the data source consistent)
    }

    def get_process_chemical_energy_per_cm2_advance_nodes(node_nm: str):
        process_steps_per_wafer = {
            '28': 566,
            '20': 642,
            '14': 670,
            '10': 812,
            '8': 1145,
            '7': 1046,
            '6': 1111,
            '5': 1089,
            '3': 1047
        }
        node_nm = str(node_nm)
        assert node_nm in process_steps_per_wafer.keys(), "Unsupported node for process chemical energy estimation."

        if node_nm in ['28', '3']:
            return PER_UNIT_PROCESS_CHEMICAL_ENERGY['logic_advanced_' + node_nm + 'nm']
        X = np.array([process_steps_per_wafer['28'], process_steps_per_wafer['3']])
        Y = np.array([PER_UNIT_PROCESS_CHEMICAL_ENERGY['logic_advanced_28nm'], PER_UNIT_PROCESS_CHEMICAL_ENERGY['logic_advanced_3nm']])
        coeffs = np.polyfit(X, Y, 1)  # Linear fit
        poly = np.poly1d(coeffs)
        return poly(process_steps_per_wafer[node_nm])
    
    # DRAM Bit Density and # process steps follows: #PS = a log10(Bit Density) + b
    # Two samples: (0.1, 200*112/94), (1, 400*168/189)
    def calc_dram_process_steps(bit_density): 
        return 118 * np.log10(bit_density) + 356


    if component_type in ['CPU', 'GPU']:
        if logic_node_nm > 28:
            process_chemical_energy = (PER_UNIT_PROCESS_CHEMICAL_ENERGY['logic_mature_lower_bound'] + PER_UNIT_PROCESS_CHEMICAL_ENERGY['logic_mature_upper_bound']) / 2 * get_equivalent_wafers(device_specs, bit_density=bit_density) # kWh
        else:
            process_chemical_energy_per_cm2 = get_process_chemical_energy_per_cm2_advance_nodes(device_specs['technology_node_nm'])
            process_chemical_energy = process_chemical_energy_per_cm2 * required_wafer_area_cm2 # kWh
    elif component_type in ['DRAM', 'HBM']:
        if bit_density is not None:
            scaling_factor = calc_dram_process_steps(bit_density) / calc_dram_process_steps(0.022) # use the estimated number of process steps based on bit density to scale the chemical energy for DRAM, with reference to the 57nm node in Boyd book
        else:
            bit_density = dram_bit_density_by_year[device_specs['production_year']-2016] if component_type == 'DRAM' else vram_bit_density.get(device_specs.get('hbm_type'), dram_bit_density_by_year[device_specs['production_year']-2016]) # use the default bit density data based on production year if customized bit density input is not provided
            scaling_factor = calc_dram_process_steps(bit_density) / calc_dram_process_steps(0.022)
        process_chemical_energy = PER_UNIT_PROCESS_CHEMICAL_ENERGY['dram'] * scaling_factor * required_wafer_area_cm2 # kWh
    else:
        # For SSD, use the same economic value based allocation, assuming that NAND impact per wafer = (value of single NAND wafer) / (value of single DRAM wafer) * DRAM impact per wafer
        assert bit_density is None, "SSD does not support customized bit density input for chemical impact estimation, please provide production_year data within the range of 2016-2024"
        year_idx = min(max(device_specs['production_year'] - 2016, 0), len(ssd_bit_density_by_year) - 1)
        reference_dram_impact_scale_factor = calc_dram_process_steps(dram_bit_density_by_year[year_idx]) / calc_dram_process_steps(0.022)
        dram_to_ssd_value_ratio = (hynix_production_data['nand_revenue_ratio'][year_idx] / hynix_production_data['nand_k_wafers'][year_idx]) / (hynix_production_data['dram_revenue_ratio'][year_idx] / hynix_production_data['estimated_dram_k_wafers'][year_idx])
        scaling_factor = reference_dram_impact_scale_factor * dram_to_ssd_value_ratio 
        process_chemical_energy = PER_UNIT_PROCESS_CHEMICAL_ENERGY['dram'] * scaling_factor * required_wafer_area_cm2 # kWh


    impacts['chemicals'] = calculate_operational_impacts_given_energy(process_chemical_energy, device_specs['production_year'], chemical_region, spatial_awareness=spatial_awareness, excluded_pollutants=['CO2e'] if component_type in ['CPU', 'GPU'] else [])
    if chemical_transportation is not None:
        if component_type in ['CPU', 'GPU']:
            shipping_mass = get_equivalent_wafers(device_specs, bit_density=bit_density) * KG_CHEMICALS_PER_WAFER_LOGIC_45NM * (node_to_layer_masks_map[device_specs['technology_node_nm']] / node_to_layer_masks_map['45']) # scale the chemical mass based on the number of layers for different nodes
        else:
            shipping_mass = scaling_factor * get_equivalent_wafers(device_specs, bit_density=bit_density) * KG_CHEMICALS_PER_WAFER_DRAM_57NM # scale the chemical mass based on the estimated chemical energy scaling factor for DRAM and SSD
        
        transportation_impacts = calculate_transport_impact(shipping_mass, distance=chemical_transportation) 
        for impact_type in MIDPOINT_SCOPES:
            impacts['chemicals'][impact_type] += transportation_impacts.get(impact_type, 0.0)
            

    # 2,4-DB is not convertable to 1,4-DCB directly, need to rework through the CF table based on Boyd book inputs
    # After investigation, the Boyd book seems only considered the energy use impact for material procurement, rather than the cradle-to-gate impact of purchased silicon wafers and process chemicals. If background CFs are used (e.g., Ecoinvent), the impact will be higher.
    # I'm using the exact energy number in Table B.1 and airborne mercury emission factor to estimate the 1,4-DCB impact factor here.
    # Showing results based on 45/57nm node from Boyd book for logic/DRAM -- flash does not have by-stage data in the book, only aggregated results
    # convert pixel heights from a figure logscale=10 to reference impact values
    # reference_impact_per_wafer = {
    #     'silicon': {
    #         # 'AP': 1e-2 * 10 **(176/208) * 443 * 0.032,  # mol H+ eq -> kg SO2 eq per 300mm wafer
    #         'AP': 6.5*443/10.7 * 0.4 / 1000, # kWh * g/kWh / 1000 = kg SO2 eq per 300mm wafer
    #         'FEP': 0,
    #         # 'TEP': 1e-2 * 10 ** (118/184) * 443,  # kg N eq per 300mm wafer
    #         'MEP': 0, # the data source is terrestrial eutrophication (kg N via air), which is not included in the current impact categories
    #         # 'FETox': 1e-3 * 10 ** (120/156) * 443,  # kg 2,4-DB eq per 300mm wafer
    #         'TETP': 6.5*443/10.7/1000*0.0038/1000*1.38e6,  # kg 1,4-DCB eq per 300mm wafer
    #         'FETP': 6.5*443/10.7/1000*0.0038/1000*1.01, # kg 1,4-DCB eq per 300mm wafer
    #         'METP': 6.5*443/10.7/1000*0.0038/1000*240, # kg 1,4-DCB eq per 300mm wafer
    #         'POFP': 1e-3 * 10 ** (19/222) * 443  # kg NOx eq per 300mm wafer
    #     },
    #     'chemicals_logic': {
    #         # 'AP': 1e-2 * 10 **(94/208) * 443 * 0.032,  # mol H+ eq -> kg SO2 eq per 300mm wafer
    #         'AP': 4.2*443/10700 * 0.4 / 1000, # kWh * g/kWh / 1000 = kg SO2 eq per 300mm wafer
    #         'FEP': 0,
    #         # 'TEP': 1e-2 * 10 ** (83/184) * 443,  # kg N eq per 300mm wafer
    #         'MEP': 0, 
    #         # 'FETox': 1e-3 * 10 ** (93/156) * 443,  # kg 2,4-DB eq per 300mm wafer
    #         'TETP': 4.2*443/10700*0.0038/1000*1.38e6,  # kg 1,4-DCB eq per 300mm wafer
    #         'FETP': 4.2*443/10700*0.0038/1000*1.01, # kg 1,4-DCB eq per 300mm wafer
    #         'METP': 4.2*443/10700*0.0038/1000*240, # kg 1,4-DCB eq per 300mm wafer
    #         'POFP': 1e-4 * 10 ** (201/222) * 443  # kg NOx eq per 300mm wafer
    #     },
    #     'chemicals_dram': {
    #         # 0.022 Gb/mm^2 is the correct bit density to use for 57nm DRAM node around the year 2008 based on the IEDM 2023 paper Figure 12, the Boyd book itself's 0.21 Gb/mm^2 in Table 6.2 is nonsense
    #         'AP': 1e-4 * 10 **(58/270) / 16 * 0.022 * EFFECTIVE_WAFER_AREA_MM2 * 0.032,  # mol H+ eq -> kg SO2 eq per 300mm wafer
    #         'FEP': 0,
    #         # 'TEP': 1e-4 * 10 ** (210/256) / 16 * 0.022 * EFFECTIVE_WAFER_AREA_MM2,  # kg N eq per 300mm wafer
    #         'MEP': 0, 
    #         # 'FETox': 1e-3 * 10 ** (70/156) * 443,  # kg 2,4-DB eq per 300mm wafer
    #         'TETP': 1e-1*10**(240/265)/ 16 * 0.022 * EFFECTIVE_WAFER_AREA_MM2 / 12000*0.0038/1000*1.38e6,  # kg 1,4-DCB eq per 300mm wafer
    #         'FETP': 1e-1*10**(240/265)/ 16 * 0.022 * EFFECTIVE_WAFER_AREA_MM2 / 12000*0.0038/1000*1.01, # kg 1,4-DCB eq per 300mm wafer
    #         'METP': 1e-1*10**(240/265)/ 16 * 0.022 * EFFECTIVE_WAFER_AREA_MM2 / 12000*0.0038/1000*240, # kg 1,4-DCB eq per 300mm wafer
    #         'POFP': 1e-5 * 10 ** (79/320) / 16 * 0.022 * EFFECTIVE_WAFER_AREA_MM2  # kg NOx eq per 300mm wafer
    #     }
    # }


    return impacts

def calculate_manufacturing_bom_impacts(device_specs, bom_template=None, upstream_transportation=None, **kwargs):
    """
    Calculate the manufacturing BOM impacts of a device based on its specifications.
    Supported midpoint categories: AP, MEP, FETP, Smog Formation
    
    Args:
        device_specs (dict): A dictionary containing device specifications. Expected keys are:
            - 'component_type': Type of component ('CPU', 'GPU', 'SSD', 'HDD', 'DRAM')
            - 'net_weight': Mass of the component in grams 
            - 'production_year': Year of production (2016-2023)
            - Other keys required for specific component types
            
    Returns:
        dict: A dictionary containing the calculated manufacturing impacts.
    """
    component_type = device_specs['component_type']
    if component_type not in ['CPU', 'GPU']:
        raise ValueError("Unsupported component type for BOM impact calculation.")
    mass = device_specs['net_weight'] / 1000  # convert g to kg
    if bom_template is None:
        if component_type == 'CPU':
            bom_template = bom_templates['CPU']
        elif device_specs['cooling_type'] == 'air':
            bom_template = bom_templates['GPU_air_cooling']
        else:
            bom_template = bom_templates['GPU_liquid_cooling']
    
    impacts = {midpoint: 0.0 for midpoint in MIDPOINT_SCOPES}
    for material, fraction in bom_template.items():
        material_mass = mass * fraction  # kg
        for midpoint in MIDPOINT_SCOPES:
            impacts[midpoint] += _material_cf_contribution(material, midpoint, material_mass)
    
    if upstream_transportation is not None:
        transportation_impacts = calculate_transport_impact(mass, distance=upstream_transportation, mass_unit='kg')
        for impact_type in MIDPOINT_SCOPES:
            impacts[impact_type] += transportation_impacts.get(impact_type, 0.0)

    return impacts

def calculate_packaging_material_impacts(packaging_mass, mass_unit='kg', paper_to_pe_ratio=3.9375):
    """
    Calculate the impacts of packaging material based on its mass and composition.
    Args:
        packaging_mass (float): The mass of the packaging material.
        mass_unit (str): The unit of the mass ('g', 'kg', 't'). Default is 'kg'.
        paper_to_pe_ratio (float): The mass ratio of paper to PE in the packaging material. Default is 3.9375 based on Dell R740 LCA report for packaging material.
    Returns:
        dict: A dictionary containing the calculated impacts for the packaging material.
    """
    # Convert mass to kg
    conversion = {
        'g': 1e-3,
        'kg': 1,
        't': 1e3
    }
    
    if mass_unit not in conversion:
        raise ValueError("Mass unit must be 'g', 'kg', or 't'")
    
    mass_in_kg = packaging_mass * conversion[mass_unit]
    mass_paper = mass_in_kg * paper_to_pe_ratio / (1 + paper_to_pe_ratio)
    mass_pe = mass_in_kg / (1 + paper_to_pe_ratio)

    impacts = {k:0.0 for k in MIDPOINT_SCOPES}
    for impact_type in MIDPOINT_SCOPES:
        impacts[impact_type] = (
            _material_cf_contribution('packaging_plastic', impact_type, mass_pe)
            + _material_cf_contribution('packaging_paper', impact_type, mass_paper)
        )

    return impacts
    

def calculate_recycling_impact(mass, mass_unit='kg', pathway='recycling', PAPER_TO_PE_RATIO=3.9375, midpoint_scopes=None):
    """
    Calculate recycling impact given mass and recycling rates
    
    Parameters:
    mass: float - mass of the component
    mass_unit: str - unit of mass ('g', 'kg', 't', default 'kg')
    pathway: str - recycling pathway ('recycling', 'landfill', 'inceneration', default 'recycling')
    PAPER_TO_PE_RATIO: float - mass ratio of paper to PE in e-waste inceneration (default 3.9375 based on Dell R740 LCA report for packaging material)
    
    Returns:
    dict - impact values for each impact type
    """

    assert pathway in ['recycling', 'landfill', 'inceneration'], "Pathway must be 'recycling', 'landfill', or 'inceneration'"

    selected_midpoint_scopes = tuple(midpoint_scopes or MIDPOINT_SCOPES)
    impacts = {k:0.0 for k in selected_midpoint_scopes}
    if pathway == 'recycling':
        return impacts

    # Convert mass to kg
    conversion = {
        'g': 1e-3,
        'kg': 1,
        't': 1e3
    }
    
    if mass_unit not in conversion:
        raise ValueError("Mass unit must be 'g', 'kg', or 't'")
    
    mass_in_kg = mass * conversion[mass_unit]
    if pathway == 'landfill':
        impact_factors = cf_eol['landfill']
        for impact_type in selected_midpoint_scopes:
            impacts[impact_type] = mass_in_kg * impact_factors.get(impact_type, 0.0)
    elif pathway == 'inceneration':
        impact_factors_pe = cf_eol['energy_recycling_PE']
        impact_factors_paper = cf_eol['energy_recycling_paper']
        mass_paper = mass_in_kg * PAPER_TO_PE_RATIO / (1 + PAPER_TO_PE_RATIO)
        mass_pe = mass_in_kg / (1 + PAPER_TO_PE_RATIO)
        for impact_type in selected_midpoint_scopes:
            impacts[impact_type] = (
                mass_pe * impact_factors_pe.get(impact_type, 0.0)
                + mass_paper * impact_factors_paper.get(impact_type, 0.0)
            )
        
    
    return impacts

def calculate_device_recycling_impact(device_specs, use_region=None, eol_region=None, transport_region=None, midpoint_scopes=None, **kwargs):
    # TODO: Add Al, Cu, and Fe recovery impacts
    """
    Calculate recycling impact of a device based on its specifications.
    Assumes 100% packaging material (mixed plastic and paper) is incinerated, and the rest of the device mass is recycled/landfilled based on recycling rate.

    This is a simplified model and does not account for different materials in the device; in reality, different materials have different recycling rates and impacts.
    
    Args:
        device_specs (dict): A dictionary containing device specifications. Expected keys are:
            - 'net_weight': Mass of the component in grams
            - 'recycling_rate': Recycling rate (default 0.8232 based on Dell R740 LCA report)
        use_region (str, optional): Override the manufacture-to-use destination region.
            Only matters here when provided via ``transport_region``.
        eol_region (str, optional): Override the use-to-EoL destination region.
        transport_region (str, optional): Convenience alias for setting both
            ``use_region`` and ``eol_region`` to the same region.
    
    Returns:
        dict: A dictionary containing the calculated recycling impacts breakdown.
    """
    working_specs = _with_derived_transport_weights(device_specs)
    mass = working_specs['net_weight']
    packaging_mass = working_specs.get('gross_weight', 0) - working_specs.get('net_weight', 0)
    recycling_rate = working_specs.get('recycling_rate', 0.8232)
    transport_distances = resolve_transport_distances(
        working_specs,
        use_region=use_region,
        eol_region=eol_region,
        transport_region=transport_region,
    )

    selected_midpoint_scopes = tuple(midpoint_scopes or MIDPOINT_SCOPES)

    impacts_inceneration = calculate_recycling_impact(
        packaging_mass,
        mass_unit='g',
        pathway='inceneration',
        midpoint_scopes=selected_midpoint_scopes,
    )
    impacts_transportation = calculate_transport_impact(
        mass,
        distance=transport_distances['use_to_eol'],
        mass_unit='g',
        midpoint_scopes=selected_midpoint_scopes,
    )
    impacts_landfill = calculate_recycling_impact(
        mass*(1-recycling_rate),
        mass_unit='g',
        pathway='landfill',
        midpoint_scopes=selected_midpoint_scopes,
    )

    total_impacts = {
        midpoint: impacts_transportation.get(midpoint, 0.0)
        + impacts_landfill.get(midpoint, 0.0)
        + impacts_inceneration.get(midpoint, 0.0)
        for midpoint in selected_midpoint_scopes
    }

    return total_impacts


def midpoint_to_endpoint(
    midpoint_impact,
    perspective=DEFAULT_RECIPE_PERSPECTIVE,
    endpoint_scopes=ECOSYSTEM_ENDPOINT_SCOPES,
    location=None,
    spatial_awareness=False,
):
    """
    Convert midpoint impact to endpoint impact.
    
    Args:
        midpoint_impact (dict): A dictionary containing midpoint impacts. Expected valid keys are:
            - 'POFP': Photochemical Oxidant Formation Potential
            - 'AP': Acidification Potential
            - 'TETP': Terrestrial Ecotoxicity Potential
            - 'FEP': Freshwater Eutrophication Potential
            - 'FETP': Freshwater Ecotoxicity Potential
            - 'MEP': Marine Eutrophication Potential
            - 'METP': Marine Ecotoxicity Potential
            - 'GWP': Global Warming Potential
            - 'WC': Water Consumption
        perspective (str): ReCiPe perspective to use. Supported values include
            ``individualistic``, ``hierarchist``/``hierarchic``, and
            ``egalitarian``. Default is ``hierarchist``.
        endpoint_scopes (str | iterable[str]): Endpoint scopes to aggregate
            over. Defaults to the ecosystem scopes only:
            ``('terrestrial', 'freshwater', 'marine')``.
        location (str, optional): Operational region/country used to override
            only the water-consumption midpoint-to-endpoint factor when
            ``spatial_awareness`` is enabled.
        spatial_awareness (bool): When ``True``, allow location-specific
            endpoint conversion for water consumption only.
    
    Returns:
        endpoint_impact (dict): A dictionary containing endpoint impacts in species`yr for each midpoint.

    Note: GWP and WC are counted twice for their dual impact on both terrestrial and freshwater ecosystems.
    """
    perspective_key = RECIPE_PERSPECTIVE_ALIASES.get(str(perspective).strip().lower())
    if perspective_key is None:
        raise ValueError(
            f"Unsupported ReCiPe perspective {perspective!r}. "
            f"Available aliases: {sorted(RECIPE_PERSPECTIVE_ALIASES)}"
        )
    if perspective_key not in midpoint_to_endpoint_factors_by_perspective:
        raise ValueError(
            f"Perspective {perspective!r} is unavailable in the loaded midpoint-to-endpoint factors. "
            f"Available perspectives: {sorted(midpoint_to_endpoint_factors_by_perspective)}"
        )

    if endpoint_scopes == "all":
        selected_scopes = tuple(midpoint_to_endpoint_factors_by_perspective[perspective_key].keys())
    elif isinstance(endpoint_scopes, str):
        selected_scopes = (endpoint_scopes,)
    else:
        selected_scopes = tuple(endpoint_scopes)

    conversion_factors = midpoint_to_endpoint_factors_by_perspective[perspective_key]
    invalid_scopes = [scope for scope in selected_scopes if scope not in conversion_factors]
    if invalid_scopes:
        raise ValueError(
            f"Unsupported endpoint scopes {invalid_scopes}. "
            f"Available scopes: {sorted(conversion_factors)}"
        )

    location_specific_water_factors = _get_location_specific_water_endpoint_factors(
        location,
        perspective_key,
        spatial_awareness=spatial_awareness,
    )
    endpoint_impact = {midpoint: 0.0 for midpoint in midpoint_impact.keys()}
    for ecosystem_type in selected_scopes:
        for midpoint in midpoint_impact.keys():
            factor = conversion_factors[ecosystem_type].get(midpoint)
            if midpoint == 'WC' and ecosystem_type in location_specific_water_factors:
                factor = location_specific_water_factors[ecosystem_type]
            if factor is not None:
                endpoint_impact[midpoint] += midpoint_impact[midpoint] * factor
    
    return endpoint_impact

def calculate_total_impact(device_specs, occupy_ratio=1.0, manufacturing_only=False, spatial_awareness=False, calculate_upstream_materials=True, bom_template=None, use_region=None, eol_region=None, transport_region=None, **transportation_kwargs):
    """
    Calculate the total embodied impact of a device based on its specifications.
    
    Args:
        device_specs (dict): A dictionary containing device specifications. Expected keys are:
            - 'component_type': Type of component ('CPU', 'GPU', 'SSD', 'HDD', 'DRAM')
            - 'net_weight': Mass of the component in grams 
            - 'distance': dict with transport distances for different modes (Truck, Ship, Air)
            - 'production_year': Year of production (2016-2024)
        use_region (str, optional): Region key from
            ``regional_device_transport_routes.json`` to override the
            manufacture-to-use route while leaving EoL unchanged unless specified.
        eol_region (str, optional): Region key from
            ``regional_device_transport_routes.json`` to override the use-to-EoL
            truck distance.
        transport_region (str, optional): Convenience alias that sets both
            ``use_region`` and ``eol_region`` unless either one is passed
            explicitly.
    
    Returns:
        impact (dict): A dictionary containing the calculated impacts breakdown.
    """
    perspective = transportation_kwargs.pop('perspective', DEFAULT_RECIPE_PERSPECTIVE)
    working_specs = _with_derived_transport_weights(device_specs)

    # Calculate manufacturing impacts and preserve an internal scope 1/2/3 split.
    manufacturing_scope_breakdown = calculate_manufacturing_impact_breakdown(
        working_specs,
        spatial_awareness=spatial_awareness,
        calculate_upstream_materials=calculate_upstream_materials,
        bom_template=bom_template,
        **transportation_kwargs,
    )

    # packaging material impacts for all components
    packaging_material_impacts = calculate_packaging_material_impacts(working_specs['gross_weight'] - working_specs['net_weight'], mass_unit='g')
    _add_midpoint_impacts(manufacturing_scope_breakdown['scope3'], packaging_material_impacts)
    manufacturing_impacts = _sum_manufacturing_scope_breakdown(manufacturing_scope_breakdown)
    manufacturing_endpoint = midpoint_to_endpoint(manufacturing_impacts, perspective=perspective)
    manufacturing_scope_breakdown_payload = _build_scope_breakdown_payload(
        manufacturing_scope_breakdown,
        perspective=perspective,
    )

    if manufacturing_only:
        total_impacts = {
            'manufacturing': {
                'midpoint': {k: manufacturing_impacts[k]* occupy_ratio for k in manufacturing_impacts.keys()},
                'endpoint': {k: manufacturing_endpoint[k]* occupy_ratio for k in manufacturing_endpoint.keys()},
                'scope_breakdown': {
                    scope_name: {
                        'midpoint': {
                            midpoint: payload['midpoint'][midpoint] * occupy_ratio
                            for midpoint in payload['midpoint'].keys()
                        },
                        'endpoint': {
                            midpoint: payload['endpoint'][midpoint] * occupy_ratio
                            for midpoint in payload['endpoint'].keys()
                        },
                    }
                    for scope_name, payload in manufacturing_scope_breakdown_payload.items()
                },
            }
        }
        return total_impacts

    # Calculate transportation and recycling impacts
    if working_specs['component_type'] != 'HDD':
        transport_distances = resolve_transport_distances(
            working_specs,
            use_region=use_region,
            eol_region=eol_region,
            transport_region=transport_region,
        )
        transport_impacts = calculate_transport_impact(
            working_specs['gross_weight'],
            transport_distances['manufacture_to_use'],
            mass_unit='g',
        )
        recycling_impacts = calculate_device_recycling_impact(
            working_specs,
            use_region=use_region,
            eol_region=eol_region,
            transport_region=transport_region,
        )
    else:
        if any(region is not None for region in (use_region, eol_region, transport_region)):
            raise ValueError(
                "Regional transport overrides are not yet supported for HDD in calculate_total_impact(). "
                "HDD still uses pre-tabulated transportation/end-of-life stage impacts from hdd_impact_factors_by_stage_year.json."
            )
        hdd_stage_impact_factors = _get_hdd_stage_impact_factors_for_year(working_specs['production_year'])
        transport_impacts = {k: 0.0 for k in MIDPOINT_SCOPES}
        recycling_impacts = {k: 0.0 for k in MIDPOINT_SCOPES}
        for impact_type in MIDPOINT_SCOPES:
            if impact_type in hdd_stage_impact_factors:
                transport_impacts[impact_type] = hdd_stage_impact_factors[impact_type].get('transportation', 0.0) * working_specs['capacity']
                recycling_impacts[impact_type] = hdd_stage_impact_factors[impact_type].get('end-of-life', 0.0) * working_specs['capacity']

    # Calculate total impacts
    transport_endpoint = midpoint_to_endpoint(transport_impacts, perspective=perspective)
    recycling_endpoint = midpoint_to_endpoint(recycling_impacts, perspective=perspective)
    total_impacts = {
        'manufacturing': {
            'midpoint': {k: manufacturing_impacts[k]* occupy_ratio for k in manufacturing_impacts.keys()},
            'endpoint': {k: manufacturing_endpoint[k]* occupy_ratio for k in manufacturing_endpoint.keys()},
            'scope_breakdown': {
                scope_name: {
                    'midpoint': {
                        midpoint: payload['midpoint'][midpoint] * occupy_ratio
                        for midpoint in payload['midpoint'].keys()
                    },
                    'endpoint': {
                        midpoint: payload['endpoint'][midpoint] * occupy_ratio
                        for midpoint in payload['endpoint'].keys()
                    },
                }
                for scope_name, payload in manufacturing_scope_breakdown_payload.items()
            },
        },
        'transportation': {
            'midpoint': {k: transport_impacts[k]* occupy_ratio for k in transport_impacts.keys()},
            'endpoint': {k: transport_endpoint[k]* occupy_ratio for k in transport_endpoint.keys()}
        },
        'recycling': {
            'midpoint': {k: recycling_impacts[k]* occupy_ratio for k in recycling_impacts.keys()},
            'endpoint': {k: recycling_endpoint[k]* occupy_ratio for k in recycling_endpoint.keys()}
        },
        'total': {
            'midpoint': {
                k: (manufacturing_impacts[k] + transport_impacts[k] + recycling_impacts[k]) * occupy_ratio for k in MIDPOINT_SCOPES
            }
        }
    }
    
    # Calculate total endpoint impacts
    total_impacts['total']['endpoint'] = midpoint_to_endpoint(total_impacts['total']['midpoint'], perspective=perspective)
    
    return total_impacts


def _with_derived_transport_weights(device_specs):
    """
    Return a shallow copy of device specs with transport/recycling masses derived in
    the same way as the legacy total-impact path, without mutating shared specs.
    """
    working_specs = dict(device_specs)
    component_type = working_specs['component_type']
    if component_type in ['DRAM', 'SSD']:
        working_specs['net_weight'] = df_storage_weight_density[f"{component_type}_g/GB"].loc[str(working_specs['production_year'])] * working_specs['capacity']
        working_specs['gross_weight'] = DRAM_SSD_PACKAGING_WEIGHT_FACTOR * working_specs['net_weight']
    elif component_type == 'HDD':
        working_specs['net_weight'] = df_storage_weight_density["HDD_net_weight_g/GB"].loc[str(working_specs['production_year'])] * working_specs['capacity']
        working_specs['gross_weight'] = df_storage_weight_density["HDD_total_weight_g/GB"].loc[str(working_specs['production_year'])] * working_specs['capacity']
    elif component_type in LOGIC_COMPONENT_TYPES:
        working_specs['gross_weight'] = CPU_GPU_PACKAGING_WEIGHT_FACTOR * working_specs['net_weight']
    return working_specs

# Calculate operational impacts for different load scenarios
def calculate_operational_impacts(device_specs, load_ratio, years=5, location='US', emission_factor_year=None, spatial_awareness=False, excluded_pollutants=None, datacenter_wue=None, pue=1.2, midpoint_scopes=None):
    """
    Calculate operational impacts based on TDP and load ratio.
    If spatial awareness is enabled, use the same location-specific midpoint
    characterization overrides as ``calculate_operational_impacts_given_energy``.
    
    Args:
        device_specs: Device specifications including TDP
        load_ratio: Ratio of actual power to TDP
        years: Operating years
    """
    if device_specs['component_type'] not in ['CPU', 'GPU']:
        achieved_power = (load_ratio-0.3)/0.7 * (device_specs['peak_power']-device_specs['idle_power']) + device_specs['idle_power']
        annual_energy = achieved_power * 24 * 365 / 1000
    else:
        tdp = device_specs['TDP']  # in watts
        annual_energy = tdp * load_ratio * 24 * 365 / 1000  # kWh/year
    
    if emission_factor_year is not None:
        return calculate_operational_impacts_given_energy(
            annual_energy * years,
            emission_factor_year,
            location=location,
            spatial_awareness=spatial_awareness,
            excluded_pollutants=excluded_pollutants,
            datacenter_wue=datacenter_wue,
            pue=pue,
            midpoint_scopes=midpoint_scopes,
        )
    
    selected_midpoint_scopes = tuple(midpoint_scopes or MIDPOINT_SCOPES)
    impacts = {k: 0.0 for k in selected_midpoint_scopes}
    # Calculate year by year impacts
    for year in range(device_specs['production_year'], device_specs['production_year'] + years):
        annual_impacts = calculate_operational_impacts_given_energy(
            annual_energy,
            year,
            location=location,
            spatial_awareness=spatial_awareness,
            excluded_pollutants=excluded_pollutants,
            datacenter_wue=datacenter_wue,
            pue=pue,
            midpoint_scopes=selected_midpoint_scopes,
        )
        for impact_type in selected_midpoint_scopes:
            impacts[impact_type] += annual_impacts.get(impact_type, 0.0)
    
    return impacts

def calculate_operational_impacts_given_energy(energy, year, location='US', energy_unit='kWh',spatial_awareness=False, excluded_pollutants=None, datacenter_wue=None, pue=1.2, midpoint_scopes=None):
    """
    Calculate operational impacts based on energy consumption in kWh
    If the mode is "location-specific", use location-specific characterization factors (from ReCiPe 2016 Handbook Appendix).
    
    Args:
        energy: Energy consumption in kWh
        location: Location for emission factors
        datacenter_wue: Optional direct water usage effectiveness (L/kWh IT).
            When provided, operational water use is modeled as
            ``energy * (datacenter_wue + pue * EWF)`` instead of ``energy * EWF``.
        pue: Power usage effectiveness multiplier applied to grid-embedded water
            in the datacenter-water case. Default is 1.2.
    """
    assert energy_unit in ['kWh', 'J', 'joules', 'Joules'], "Energy unit must be 'kWh' or 'Joules'"
    if pue <= 0:
        raise ValueError("pue must be positive.")

    if energy_unit in ['J', 'joules', 'Joules']:
        energy = energy / 3.6e6  # Convert Joules to kWh
    # Initialize impacts dictionary
    selected_midpoint_scopes = tuple(midpoint_scopes or MIDPOINT_SCOPES)
    impacts = {k: 0.0 for k in selected_midpoint_scopes}
    air_pollutants = ['CO2e', 'SOx', 'NOx', 'NMVOC', 'NH3', 'airborne mercury','EWF']
    for pollutant in excluded_pollutants or []:
        if pollutant in air_pollutants:
            air_pollutants.remove(pollutant)

    def _get_emission_factor(pollutant, loc, idx):
        normalized_loc = _normalize_emission_factor_location(loc)
        if pollutant == 'EWF' and normalized_loc == 'IA' and 'MidWest' in unified_emission_factors[pollutant]:
            return unified_emission_factors[pollutant]['MidWest'][idx]
        if pollutant in ['NH3', 'airborne mercury', 'NMVOC'] and normalized_loc in ['IN', 'TX', 'VA', 'NE', 'MO', 'OH', 'WY']:
            return unified_emission_factors[pollutant]['US'][idx]
        if normalized_loc in unified_emission_factors[pollutant]:
            return unified_emission_factors[pollutant][normalized_loc][idx]

        # EDGAR keys are pollutant-specific; if a pollutant is missing there,
        # fall back to the base country mix before defaulting to US.
        edgar_parent = _get_edgar_parent_location(normalized_loc)
        if edgar_parent and edgar_parent in unified_emission_factors[pollutant]:
            return unified_emission_factors[pollutant][edgar_parent][idx]

        return unified_emission_factors[pollutant]['US'][idx]
    
    # Get emission factors for that year (use 2024 if beyond)
    year_idx = min(year, 2024) - 2016  # 2016 is index 0
    pollutant_emission_factors = {
        pollutant: _get_emission_factor(pollutant, location, year_idx) for pollutant in air_pollutants
    }
    
    # Calculate impacts with optional location-specific overrides
    location_specific_cfs = _get_location_specific_midpoint_cfs(location, spatial_awareness=spatial_awareness)

    for pollutant in air_pollutants:
        if pollutant == 'EWF' and datacenter_wue is not None:
            indirect_emission = energy * (datacenter_wue + pue * pollutant_emission_factors[pollutant]) / 1000
        else:
            indirect_emission = energy * pollutant_emission_factors[pollutant] / 1000  # convert g to kg / L to m3
        for impact_type in selected_midpoint_scopes:
            pollutant_cf = _get_midpoint_impact_factor(impact_type, pollutant, location_specific_cfs)
            if pollutant_cf is not None:
                impacts[impact_type] += pollutant_cf * indirect_emission
    

    return impacts
